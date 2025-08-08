import pandas as pd
import numpy as np
import nanoid
import random
from difflib import unified_diff

from langflow.schema.dataframe import DataFrame, Data
from langflow.base.forecasting_common.models.forecast_data_model import ForecastDataModel
from langflow.base.forecasting_common.constants import FORECAST_INT_TO_SHORT_MONTH_NAME, ForecastModelInputTypes, ForecastModelTimescale
from langflow.base.forecasting_common.models.forecast_meta_data import (ForecastMetaDataSeries, 
                                                                        ForecastMetaDataFrame,
                                                                        ForecastMetaDataRange,
                                                                        ForecastMetaDataSeriesSchema, 
                                                                        ForecastMetaDataFrameSchema,
                                                                        ForecastMetaDataRangeSchema,
                                                                        ForecastDataSeriesMetaDataStepTypes, 
                                                                        ForecastDataSeriesMetaDataAction, 
                                                                        ForecastDataSeriesMetaDataDataType, 
                                                                        ForecastDataSeriesMetaDataValidationSchema, 
                                                                        ForecastDataSeriesMetaDataValidateInputRestrictions,
                                                                        ForecastMetaDataSeriesIdGenerator)

from langflow.base.forecasting_common.builders.excel.forecast_excel_iterator import ForecastPredIterator
from langflow.base.forecasting_common.builders.forecast_builder_excel_TB import (IdToCellReferenceMap,
                                                                                 IdToCellReferenceMaps,
                                                                                 ForecastPredRef)

from openpyxl import Workbook, worksheet, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection
import pickle


#test configuration
MODE = "GENERATE"
#MODE = "TEST"
LOCATION = "output/"

TEST_NAME = "forecast_meta_data"
INPUT_FILE = "_input.pickle"
EXPECTED_OUTPUT_FILE = "_expected_output.pickle"



def generate_pred(list_of_sheets, sheets_to_generate, sheet_tracker, list_of_special_address_types, default_object_id, list_of_object_ids, objects_to_generate, num_elements_in_row,
                  max_shift, address_map):
    # generate a row id
    row_id = "Row_" + nanoid.generate(size=5)

    # randomly determine the parameters for this pred
    tab_name = list_of_sheets[random.randint(0, sheets_to_generate-1)]           # tab name

    #add the row to the mapping tables
    key_id = list(sheet_tracker[tab_name].keys())[-1]
    latest_cell_ref = sheet_tracker[tab_name][key_id].offset(row = 1, column = 0) # get cell on the next available row on sheet

    # determine the address for the pred ref
    attributes = list_of_special_address_types[random.randint(1, len(list_of_special_address_types)-1)]
    cell_ref_str = ""

    match attributes:
        case "rel":
            ref_object = default_object_id
            cell_ref_str = row_id
        
        case "full":
            ref_object = list_of_object_ids[random.randint(0, objects_to_generate-1)]
            cell_ref_str = f"{ref_object}.{row_id}"

        case "indiv":
            ref_object = default_object_id
            indiv_index = random.randint(0, num_elements_in_row-1)
            cell_ref_str = f"{row_id}:{indiv_index}"

        case "shift":
            ref_object = default_object_id
            shift_num = random.randint(1, max_shift)
            cell_ref_str = f"{row_id}[{-shift_num}]"

        case "full/indiv":
            ref_object = list_of_object_ids[random.randint(0, objects_to_generate-1)]
            indiv_index = random.randint(0, num_elements_in_row-1)
            cell_ref_str = f"{ref_object}.{row_id}:{indiv_index}"

        case "full/shift":
            ref_object = list_of_object_ids[random.randint(0, objects_to_generate-1)]
            shift_num = random.randint(1, max_shift)
            cell_ref_str = f"{ref_object}.{row_id}[{-shift_num}]"

        case "rel/indiv/shift":
            ref_object = default_object_id
            indiv_index = random.randint(0, num_elements_in_row-1)
            shift_num = random.randint(1, max_shift)
            cell_ref_str = f"{row_id}:{indiv_index}[{-shift_num}]"

        case "full/indiv/shift":
            ref_object = list_of_object_ids[random.randint(0, objects_to_generate-1)]
            indiv_index = random.randint(0, num_elements_in_row-1)
            shift_num = random.randint(1, max_shift)
            cell_ref_str = f"{ref_object}.{row_id}:{indiv_index}[{-shift_num}]"


    sheet_tracker[tab_name][f"{ref_object}_{row_id}"] = latest_cell_ref
    address_map.add(id = row_id, tab_name = tab_name, cell_ref = latest_cell_ref, ref_map_id = ref_object)

    return(sheet_tracker, address_map, cell_ref_str)





def generate_preds(num_preds, list_of_sheets, sheets_to_generate, sheet_tracker, list_of_special_address_types, default_object_id, list_of_object_ids, objects_to_generate, num_elements_in_row, max_shift, address_map):

    # PRED SECTION
    # ------------
    list_of_preds = []
    #num_preds = random.randint(1, preds_to_generate)    # must have at least one pred
    print(f"\tGenerate {num_preds} preds")

    for j in range(num_preds):
        (sheet_tracker, address_map, cell_ref_str) = generate_pred(list_of_sheets, sheets_to_generate, sheet_tracker, list_of_special_address_types, default_object_id, list_of_object_ids, objects_to_generate, num_elements_in_row, max_shift, address_map)
        list_of_preds.append(cell_ref_str)

    return(sheet_tracker, address_map, list_of_preds)





def print_header(col: ForecastMetaDataSeries, num_elements_in_row: int) -> str:
    header = ""

    #if (ForecastMetaDataSeriesSchema.RANGES not in col.meta_data.keys()) or (col.meta_data[ForecastMetaDataSeriesSchema.RANGES] is None):
    if(col.has_ranges()):
        header += f"RANGES:\tNUM_RANGES: {len(col.meta_data[ForecastMetaDataSeriesSchema.RANGES])}"
    else:
        header += f"NO RANGE:\tCOUNT: None ({num_elements_in_row})\tPREDS: {col.meta_data[ForecastMetaDataSeriesSchema.PRED]}"
    
    return(header)




def print_pred_list(list_of_preds: list[str], address_map: IdToCellReferenceMaps) -> str:
    pred_output = ""

    default_card = address_map.default_ref_map_id

    for pred in list_of_preds:
        (full_id, rel_id, single_value, shift_value, has_full_id, has_single_value, has_shift_value) = ForecastMetaDataSeriesIdGenerator.parse_id(id = pred, default_full_id = default_card)
        pred_output += f"[{pred}: {has_full_id}, {has_single_value}, {has_shift_value}] "



    return(pred_output)




# ============
# MAIN PROGRAM
# ============


def main():

    input_file_name = LOCATION + TEST_NAME + INPUT_FILE
    output_file_name = LOCATION + TEST_NAME + EXPECTED_OUTPUT_FILE

    # create some sheets
    wb = Workbook()

    sheets_to_generate = 3
    list_of_sheets = [f"Sheet_{i}"  for i in range(sheets_to_generate)]
    sheet_tracker = {}


    for sheet in list_of_sheets:
        wb.create_sheet(sheet)
        sheet_tracker[sheet] = {}
        sheet_tracker[sheet]["top"] = wb[sheet].cell(row = 1, column = 1)


    # ==================
    # GENERATE TEST DATA
    # ==================
    if MODE == "GENERATE":

        num_elements_in_row = 10
        objects_to_generate = 5
        ranges_to_generate = 4
        preds_to_generate = 4
        actions_to_generate = 20
        max_shift = 3

        # mapping tables
        list_of_special_address_types = ["rel", "full", "indiv", "shift", "full/indiv", "full/shift", "rel/indiv/shift", "full/indiv/shift"]
        #list_of_sheets = [f"Sheet_{i}"  for i in range(sheets_to_generate)]
        list_of_object_ids = ["Component_" + nanoid.generate(size=5) for id in range(objects_to_generate)]
        list_of_ranges = []

        default_object_id = list_of_object_ids[0]

        # create some objects which will hold maps
        address_map = IdToCellReferenceMaps(default_ref_map_id = default_object_id, default_num_elements = num_elements_in_row)

        for i in range(1, objects_to_generate):
            address_map.create_ref_map(ref_map_id = list_of_object_ids[i], default_num_elements = num_elements_in_row)


        # ACTION SECTION
        # --------------
        list_of_actions = {}
        
        for j in range(actions_to_generate):
            num_preds = random.randint(1, preds_to_generate)    # must have at least one pred

            list_of_ranges = []
            count_left = num_elements_in_row

            print(f"\nNew action")
            id_for_action = f"Row_{nanoid.generate(size=5)}"


            # RANGE SECTION
            # -------------
            num_ranges = random.randint(0,ranges_to_generate)
            print(f"Generate {num_ranges} ranges")

            if(num_ranges == 0):
                list_of_ranges = None
                (sheet_tracker, address_map, list_of_preds) = generate_preds(num_preds, list_of_sheets, sheets_to_generate, sheet_tracker, list_of_special_address_types, default_object_id, list_of_object_ids, objects_to_generate, num_elements_in_row, max_shift, address_map)

            else:
                for i in range(num_ranges):
                    if(i == num_ranges-1):
                        count = None
                    else:
                        # randomize the count of elements for the range
                        count = random.randint(1, count_left-(num_ranges-i))
                        count_left -= count

                    # create a range and add to list
                    (sheet_tracker, address_map, list_of_preds) = generate_preds(num_preds, list_of_sheets, sheets_to_generate, sheet_tracker, list_of_special_address_types, default_object_id, list_of_object_ids, objects_to_generate, num_elements_in_row, max_shift, address_map)
                    list_of_ranges.append(ForecastMetaDataRange(count = count, pred = list_of_preds, args = None, objs = None))

            # add list_of_rages to an action
            if list_of_ranges is None:

                # add action to list of actions
                list_of_actions[id_for_action] = ForecastMetaDataSeries(id = id_for_action,
                                                                        step_type = ForecastDataSeriesMetaDataStepTypes.EPIDEMIOLOGY,
                                                                        data_type = ForecastDataSeriesMetaDataDataType.FLOAT,
                                                                        display_type = ForecastDataSeriesMetaDataDataType.INT,
                                                                        validation = [{ForecastDataSeriesMetaDataValidationSchema.INPUT_RESTRICTION: ForecastDataSeriesMetaDataValidateInputRestrictions.READ_ONLY}],
                                                                        pred = list_of_preds,)
            else:
                # add action to list of actions
                list_of_actions[id_for_action] = ForecastMetaDataSeries(id = id_for_action,
                                                                        step_type = ForecastDataSeriesMetaDataStepTypes.EPIDEMIOLOGY,
                                                                        action = ForecastDataSeriesMetaDataAction.SUM,
                                                                        ranges = list_of_ranges,
                                                                        data_type = ForecastDataSeriesMetaDataDataType.FLOAT,
                                                                        display_type = ForecastDataSeriesMetaDataDataType.INT,
                                                                        validation = [{ForecastDataSeriesMetaDataValidationSchema.INPUT_RESTRICTION: ForecastDataSeriesMetaDataValidateInputRestrictions.READ_ONLY}],)

        model_to_test = ForecastMetaDataFrame(id = "ABC",
                                            INPUT_TYPE = ForecastModelInputTypes.TIME_BASED,
                                            start_year = 2027,
                                            start_month = 1,
                                            num_periods = 5,
                                            model = list_of_actions)
        json_output = model_to_test.to_json()
        print(json_output)

        

        # SAVE OBJECTS TO PICKLE FILE
        inputs_to_save = {"model_to_test": model_to_test,
                          "num_elements_in_row": num_elements_in_row,
                          "address_map": address_map,
                          "sheet_tracker": sheet_tracker,
                          "json_output": json_output}
        
        with open(input_file_name, "wb") as file_handle:
            pickle.dump(inputs_to_save, file_handle)





    # ==================
    # OR, LOAD TEST DATA
    # ==================

    else:
        with open(input_file_name, "rb") as file_handle:
            inputs_to_load = pickle.load(file_handle)

        model_to_test = inputs_to_load["model_to_test"]
        num_elements_in_row = inputs_to_load["num_elements_in_row"]
        address_map = inputs_to_load["address_map"]
        sheet_tracker = inputs_to_load["sheet_tracker"]
        json_output = inputs_to_load["json_output"]

        with open(output_file_name, "rb") as file_handle:
            outputs_to_load = pickle.load(file_handle)
        
        original_output_results = outputs_to_load["output_results"]



    # ============
    # RUN THE TEST
    # ============
    output_results = ""

    for col in model_to_test.model.keys():

        output_results += "\n\n"
        output_results = output_results +"\n" + print_header(model_to_test.model[col], num_elements_in_row)
        forecast_predictor = ForecastPredIterator(col = model_to_test.model[col],
                                                address_maps = address_map,
                                                default_card = address_map.default_ref_map_id,
                                                total_elements = num_elements_in_row)
    
        i = 0
        left_in_range = 0

        if model_to_test.model[col].has_ranges():
            left_in_range = 0
            j = 0

            for pred_values in forecast_predictor:
                ranges = model_to_test.model[col].meta_data[ForecastMetaDataSeriesSchema.RANGES]
                left_in_range -= 1

                if(left_in_range < 0):
                    next_range = ranges[j]

                    if(next_range.meta_data[ForecastMetaDataRangeSchema.COUNT] is None):
                        left_in_range = num_elements_in_row - i - 1
                    else:
                        left_in_range = next_range.meta_data[ForecastMetaDataRangeSchema.COUNT] - 1

                    output_results += f"\n\nRange: {j}\tcount: {next_range.meta_data[ForecastMetaDataRangeSchema.COUNT]}\t{print_pred_list(next_range.meta_data[ForecastMetaDataRangeSchema.PRED], address_map)}"
                    j += 1


                output_results += f"\n{i} {j-1} {left_in_range}.  {pred_values}"
                i += 1
            
        else:
            for pred_values in forecast_predictor:
                output_results += f"\n{i}.  {pred_values}"
                i += 1

        print(output_results)

    
    # =====================
    # SAVE THE TEST RESULTS
    # =====================

    if MODE == "GENERATE":
        # SAVE OBJECTS TO PICKLE FILE
        outputs_to_save = {"output_results": output_results}
        
        with open(output_file_name, "wb") as file_handle:
            pickle.dump(outputs_to_save, file_handle)


    # ====================
    # ANALYZE TEST RESULTS
    # ====================

    else:
        if(original_output_results == output_results):
            print("TEST SUCCESSFUL")
        else:
            print("TEST FAILED")
            print(original_output_results)
            diff = unified_diff(original_output_results.splitlines(), output_results.splitlines(), lineterm='')
            print("\n\nDIFF ANALYSIS:\n".join(list(diff)))









if __name__ == "__main__":
    main()
