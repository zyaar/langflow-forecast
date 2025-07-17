#####################################################################
# forecast_renderer_excel_TB.py
#
# Implements the a summation component.  It's already implemented everywhere
# this just makes it explicit (for visual presentation purposes)
# 
# INPUTS:  DataFrame
# OUTPUTS:  DataFrame
#
#####################################################################


from typing import List, Dict, Tuple, Any
from datetime import datetime
import pandas as pd
import numpy as np
from langflow.schema.dataframe import DataFrame, Data


# FORECAST SPECIFIC IMPORTS
# =========================
from langflow.base.forecasting_common.constants import FORECAST_INT_TO_SHORT_MONTH_NAME, ForecastModelInputTypes, ForecastModelTimescale
from langflow.base.forecasting_common.models.forecast_data_model import ForecastDataModel
from langflow.base.forecasting_common.models.forecast_meta_data import (ForecastMetaDataSeries, 
                                                                        ForecastMetaDataFrame, 
                                                                        ForecastMetaDataSeriesSchema, 
                                                                        ForecastMetaDataFrameSchema,
                                                                        ForecastDataSeriesMetaDataStepTypes, 
                                                                        ForecastDataSeriesMetaDataAction, 
                                                                        ForecastDataSeriesMetaDataDataType, 
                                                                        ForecastDataSeriesMetaDataValidationSchema, 
                                                                        ForecastDataSeriesMetaDataValidateInputRestrictions,
                                                                        ForecastDataSeriesMetaDataComparisonType)


# COMPONENT SPECIFIC IMPORTS
# ==========================
from datetime import datetime
from enum import Enum
import shutil
from openpyxl import Workbook, worksheet, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection

from langflow.base.forecasting_common.renderers.excel.forecast_excel_base_helpers import ForecastExcelBaseHelpers
from langflow.base.forecasting_common.renderers.excel.forecast_excel_validation_builder import ForecastExcelValidationRuleBuilder
from langflow.base.forecasting_common.renderers.excel.forecast_excel_cell_style_builder import ForecastExcelCellStyleBuilder



# CONFIG STUFF
FORECAST_EXCEL_PROTECT_WORKSHEET = False


# CLASSES
# =======


# Enum of EXCEL_ARITHMETIC_FUNCTIONS
class ForecastBuilderExcelArithmeticFunctions(str, Enum):
    ADD = " + "
    SUB = " - "
    PROD = " * "
    DIV = " / "



# IdToCellReferenceMap
# map ForecastMetaDataFrame id's to the cell locations that they represent
class IdToCellReferenceMap():
    # INSTANCE VARIABLES
    # id_to_ref_map - a dictionary which maps all ForecastDataModel IDs to the tab and the cell reference of their rows in excel

    # __init__ function, does nothing right now
    id_to_ref_map = {}

    def __init__(self):
        pass

    # Add a new entry to the map: id is the key, tab_name and cell references are the values
    # TODO:  change the returned value of 'Any' into the name of the 'cell' object
    def add(self, id: str, tab_name: str, cell_ref: Cell):
        if(id in self.id_to_ref_map.keys()):
            raise ValueError(f"\n* IdToCellReferenceMap.add: error, id {id} already exists in map:\n{self.id_to_ref_map.keys()}")

        self.id_to_ref_map[id] = {"tab": tab_name, "ref": cell_ref}


    # Given an id, return the: tab_name, cell object pointing to the start of row
    # TODO:  change the returned value of 'Any' into the name of the 'cell' object
    def get(self, id: str) -> Tuple[str, Any]:
        try:
            full_ref = self.id_to_ref_map[id]
            tab_name = full_ref["tab"]
            cell_ref = full_ref["ref"]
        except:
            raise ValueError(f"\n* IdToCellReferenceMap.get: error, id {id} not found in map:\n{self.id_to_ref_map.keys()}")
        
        return(tab_name, cell_ref)
    


        

# ForecastRendererExcel
# Class which renders a Forecast Model player for excel using an Time Based model
class ForecastRendererExcelTB():
    # CONSTANTS
    # =========

    # FORECAST DATA MODEL CONSTANTS
    DATAMODEL_ACTION_COL = "action"
    DATAMODEL_PRED_COL = "input_rows"


    # EXCEL PLAYER CONSTANTS
    EXCEL_REQUIRED_WORKBOOK_TABS = ["Summary"]

    # PLAYER ROW COLUMN LAYOUT  (NOTE:  per openpyxl, all ROW colum numbers are 1's indexed, not 0's indexed like regular python)
    EXCEL_START_ROW = 4 # The first row of a worsheet for any rendering, 
    EXCEL_START_COL = 2 # The first row of a worsheet for any rendering
    EXCEL_LABEL_COL = EXCEL_START_COL               # Row label
    EXCEL_ID_COL = EXCEL_LABEL_COL+1                # ForecastMetaDataSeries ID for row
    EXCEL_NAME_COL = EXCEL_ID_COL+1                 # User doing the data entry's name
    EXCEL_COUNTRY_COL = EXCEL_NAME_COL+1            # Country
    EXCEL_PRODUCT_COL = EXCEL_COUNTRY_COL+1         # Product
    EXCEL_INDICATION_COL = EXCEL_PRODUCT_COL+1      # Indication for product (if needed)
    EXCEL_VALUES_START_COL = 4 # NOTE:  All COLUMN references are 1's indexed, not zero-indexed per openpyxl


    # VARIABLES
    # =========

    # CLASS VARIABLES
    input_type = ForecastModelInputTypes.TIME_BASED

    # data validation objects
    dv_integer = DataValidation(type="whole") # ensure it's an integer
    dv_float = "" # ensure it's a float (i.e. numeric)
    dv_currency = "" # ensure it's currency
    dv_percent = "" # ensure it's a percent

    dv_lt = ""  # less than validation rule
    dv_le = ""  # less than or equal to rule
    dv_eq = ""  # equal to rule
    dv_ge = ""  # greater to equal to rule
    dv_gt = ""  # greater than rule



    # INSTANCE VARIABLES
    # ------------------
    # start_year - start year of forecast
    # start_month - start of fiscal year (or 1 - January for calendar year)
    # timescale - the minimum unit for a period in the forecast (can be MONTH or YEAR)
    # num_periods - the number of periods in the forecast
    # forecast_model - the forecast model generated by the DESIGNER
    # output_location - the location to put the rendered player
    # hasTemplate - True if the render uses a template, False if not
    # template_location (optional) - the location to load the template
    # template (optional) - a template to use when developing the player
    # template_num_sheets - the number of sheets that came with the template (all our sheets must go in front of those)

    # player_model - the object model for the player being developed (openpyxl.Workbook class)
    # row_trackers - dict of ints, each tab in the spreadsheet is the key, and the current location to add a new row is the int
    # id_cellref_map - holds the mapping from all IDs to excel cell references (full references including tab name and cell coordinates)



    # CONSTRUCTOR
    # ===========
    # The constructor gets all the variables which are specific to this particular renderer vs generic to all implementers of RENDER INTERFACE.
    # In the case of this excel renderer, this include if we are using a template and where that template is located, as well as the location to
    # save the output file to
    #
    # INPUTS:
    #   output_location - location to save the output player
    #   template_location (optional) - if we are using a template file, where is the file located
    # 
    # OUTPUTS:
    #   N/A

    def __init__(self,
                 data_frame: DataFrame | pd.DataFrame,
                 meta_data:  ForecastMetaDataFrame,
                 output_location: str,
                 template_location: str = None):

        self.data_frame = data_frame
        self.meta_data = meta_data
        
        self.output_location = output_location

        if(template_location is None):
            self.hasTemplate = False
            self.template_location = None
        else:
            self.hasTemplate = True
            self.template_location = template_location



    
    # BUILDER INTERFACE FUNCTIONS
    # ==========================

    # render_player
    # RENDER INTERFACE:  render a player and return to the caller function 
    #  
    # INPUTS:
    #   TBD
    # 
    # OUTPUTS:
    #   TBD
    def render_player(self):

        # save these as instance variables
        self.start_year = self.meta_data.meta_data[ForecastMetaDataFrameSchema.START_YEAR]
        self.start_month = self.meta_data.meta_data[ForecastMetaDataFrameSchema.START_MONTH]
        self.num_periods = self.meta_data.meta_data[ForecastMetaDataFrameSchema.NUM_PERIODS]
        self.timescale = self.meta_data.meta_data[ForecastMetaDataFrameSchema.TIMESCALE]
        self.input_type = self.meta_data.meta_data[ForecastMetaDataFrameSchema.INPUT_TYPE]

        # initial setup when rendering
        self._initialize_new_render()

        # generate the model
        self._build_model_excel()

        # save the model
        self._finalize_new_render()



    # action_DATES
    # RENDER INTERFACE:  handle the DATES action by creating a row of dates
    #  
    # INPUTS:
    #   id = the id of the row
    # 
    # OUTPUTS:
    #   NA
    def action_DATES(self, id: str):
        curr_meta_col = self.meta_data.model[id] # get current MetaDataSeries col
        curr_df_col = self.data_frame[id] # get current DataFrame col

        # dates are in strings formatted as: '2027-12-31T00:00:00', so convert to datetime
        # so that openpyxl can correctly pass them to excel
        values = [datetime.strptime(s, "%Y-%m-%dT%H:%M:%S") for s in curr_df_col.to_list()]

        self._add_values_row(id = id, tab_name = self.EXCEL_REQUIRED_WORKBOOK_TABS[0], values = values, curr_row_meta_data = curr_meta_col, row_name = curr_meta_col.meta_data[ForecastMetaDataSeriesSchema.DISPLAY_NAME])
        


    # action_INPUT
    # RENDER INTERFACE:  handle the INPUT action by putting a row of values and blank cells for entering data
    #  
    # INPUTS:
    #   id = the id of the row
    # 
    # OUTPUTS:
    #   NA
    def action_INPUT(self, id: str):
        curr_meta_col = self.meta_data.model[id] # get current MetaDataSeries col
        curr_df_col = self.data_frame[id] # get current DataFrame col

        self._add_values_row(id = id, tab_name = self.EXCEL_REQUIRED_WORKBOOK_TABS[0], values = curr_df_col.to_list(), curr_row_meta_data = curr_meta_col, row_name = curr_meta_col.meta_data[ForecastMetaDataSeriesSchema.DISPLAY_NAME])
    


    # action_SUM
    # RENDER INTERFACE:  handle the SUM action by creating a row SUM'd variables based on the PREDs provided
    #  
    # INPUTS:
    #   id = the id of the row
    # 
    # OUTPUTS:
    #   NA
    def action_SUM(self, id: str):
        curr_meta_col = self.meta_data.model[id] # get current MetaDataSeries col
        curr_df_col = self.data_frame[id] # get current DataFrame col

        self._add_arith_row(arith_funct = ForecastBuilderExcelArithmeticFunctions.ADD,
                            id = id,
                            tab_name = self.EXCEL_REQUIRED_WORKBOOK_TABS[0], 
                            values = curr_df_col.to_list(),
                            curr_row_meta_data = curr_meta_col,
                            row_name = curr_meta_col.meta_data[ForecastMetaDataSeriesSchema.DISPLAY_NAME])




    # action_PROD
    # RENDER INTERFACE:  handle the PROD action by creating a row PROD'd variables based on the PREDs provided
    #  
    # INPUTS:
    #   id = the id of the row
    # 
    # OUTPUTS:
    #   NA
    def action_PROD(self, id: str):
        curr_meta_col = self.meta_data.model[id] # get current MetaDataSeries col
        curr_df_col = self.data_frame[id] # get current DataFrame col

        self._add_arith_row(arith_funct = ForecastBuilderExcelArithmeticFunctions.PROD,
                            id = id,
                            tab_name = self.EXCEL_REQUIRED_WORKBOOK_TABS[0], 
                            values = curr_df_col.to_list(),
                            curr_row_meta_data = curr_meta_col,
                            row_name = curr_meta_col.meta_data[ForecastMetaDataSeriesSchema.DISPLAY_NAME])




    # action_SUB
    # RENDER INTERFACE:  handle the subtraction action by creating a row of subtracted variables based on the PREDs provided
    #  
    # INPUTS:
    #   id = the id of the row
    # 
    # OUTPUTS:
    #   NA
    def action_SUB(self, id: str):
        curr_meta_col = self.meta_data.model[id] # get current MetaDataSeries col
        curr_df_col = self.data_frame[id] # get current DataFrame col

        self._add_arith_row(arith_funct = ForecastBuilderExcelArithmeticFunctions.SUB,
                            id = id,
                            tab_name = self.EXCEL_REQUIRED_WORKBOOK_TABS[0], 
                            values = curr_df_col.to_list(),
                            curr_row_meta_data = curr_meta_col,
                            row_name = curr_meta_col.meta_data[ForecastMetaDataSeriesSchema.DISPLAY_NAME])




    # BUILDER INTERNAL FUNCTIONS
    # ==========================

    # _initialize_new_render
    # All the excel specific set-up steps to start the process of building a new player:
    # copy the template, open a new workbook object, set-up the initial tabs, 
    #  set-up the id to cell references tracking structures, etc.
    #  
    # INPUTS:
    #   TBD
    # 
    # OUTPUTS:
    #   TBD
    def _initialize_new_render(self):
        # create new workbook or copy template
        if(self.hasTemplate):
            # since openpyxl has no object copy method for a workbook, we have to copy the template workbook
            # to the new filename and then open that
            shutil.copyfile(self.template_location, self.output_location)
            self.player_model = load_workbook(self.output_location)

        else:
            self.player_model = Workbook()

            # if any worksheets were created in the default object, get rid of them
            num_sheets = len(self.player_model.sheetnames)

            if(num_sheets > 0):
                for i in range(num_sheets):
                    ws = ForecastExcelBaseHelpers.get_ws(0, self.player_model)
                    self.player_model.remove(ws)
                    
        
        # generate the core tabs if they don't exist
        self._generate_core_tabs()

        # initialize the current row trackers
        self.row_trackers = {}
        for tab in self.EXCEL_REQUIRED_WORKBOOK_TABS:
            self.row_trackers[tab] = self.EXCEL_START_ROW    # NOTE:  All row references are 1-based based per openpyxl

        # set up the mapper between ForecastDataModel id's and cell references in excel
        self.id_cellref_map = IdToCellReferenceMap()

        

    # _finalize_new_render
    # All the excel specific steps to finalize the new player, mostly saving it to a target file location
    # TODO:  extend this to add save to sharepoint functionality
    #  
    # INPUTS:
    #   TBD
    # 
    # OUTPUTS:
    #   TBD
    def _finalize_new_render(self):
        self.player_model.save(self.output_location)

        # TODO:  figure out what the return here, the intent is to provide a way to
        # return things in realtime if a renderer doesn't save to a file (for example a dynamic JSON file)
        # but have to figure out what (if anything) to return when it does save a file
        return(self.player_model)


    
    # _build_model_excel
    # The main dispathcher for the builder, iterates over all the rows in the ForecastMetaDataFrame model, and dispatches
    # different functions the handle the different ACTIONS to build the excel player
    #  
    # INPUTS:
    #   NA
    # 
    # OUTPUTS:
    #   NA
    def _build_model_excel(self):
        model = self.meta_data.model

        # iterate through the forecast model, column by column, dispatching as needed
        for id in model:
            # get the next ForecastMetaDataSeries in the model
            curr_col = model[id] # get the current ForecastMetaDataSeries

            # dispatch based on the action
            match curr_col.meta_data[ForecastMetaDataSeriesSchema.ACTION]:
                case ForecastDataSeriesMetaDataAction.DATES:
                    self.action_DATES(id)

                case ForecastDataSeriesMetaDataAction.INPUT:
                    self.action_INPUT(id)

                case ForecastDataSeriesMetaDataAction.SUM:
                    self.action_SUM(id)

                case ForecastDataSeriesMetaDataAction.PROD:
                    self.action_PROD(id)

                case ForecastDataSeriesMetaDataAction.SUB:
                    self.action_SUB(id)

                case ForecastDataSeriesMetaDataAction.STEP_INIT:
                    pass

                case _:
                    raise ValueError(f"\n*  _build_model_excel:  Unknown action {curr_col.meta_data[ForecastMetaDataSeriesSchema.ACTION]}")
                


    # _build_validation_excel
    # The validation rules dispathcher for the builder, takes the current cell and iterates over all the elements in a list of 
    # validation elements rules, dispatching the different functions the handle the different validation needs of that cell
    #  
    # INPUTS:
    #   curr_cell - openpyxl Cell object, the current cell being working on
    # 
    # OUTPUTS:
    #   NA

    def _build_validation_excel(self, 
                                curr_cell_meta_data: ForecastMetaDataSeries, 
                                curr_cell: Cell,
                                data_type: ForecastMetaDataSeriesSchema, 
                                display_type: ForecastMetaDataSeriesSchema, 
                                validation_rules: List[Dict[ForecastDataSeriesMetaDataValidationSchema, Any]],
                                curr_formula: str = None,
                                apply_to_preds: bool = False):
        
        if ((apply_to_preds) and (curr_cell_meta_data.meta_data[ForecastMetaDataSeriesSchema.PRED] is None)):
            raise ValueError(f"\n*  _build_validation_excel:  apply_to_preds is True, but no pred field in ForecastDataSeries meta_data")
        
        for rule in validation_rules:
            rule_type = list(rule.keys())[0]
            rule_value = rule[rule_type]

            match rule_type:
                case ForecastDataSeriesMetaDataValidationSchema.INPUT_RESTRICTION:
                    self._add_input_restrictions(curr_cell = curr_cell,
                                                 data_type = data_type,
                                                 display_type = display_type,
                                                 restriction = rule_value,
                                                 curr_cell_meta_data = curr_cell_meta_data)

                case ForecastDataSeriesMetaDataValidationSchema.VALUE_CHECK:
                    # if preds exists, they needed to determine where to put the value check, so generate the formulas
                    # since data_validation in excel CANNOT work with Worksheet names (i.e. A1 works, 'Summary'!A1 does not)
                    # we generate the values without ws names
                    preds = curr_cell_meta_data.meta_data[ForecastMetaDataSeriesSchema.PRED]

                    if apply_to_preds:
                        preds = self._ids_to_formula_refs(preds, col_num = curr_cell.column, with_ws_name = False) 

                    # adds the value comparison
                    self._add_value_comparison(curr_cell = curr_cell,
                                               comparison_type = rule_value,
                                               curr_cell_meta_data = curr_cell_meta_data,
                                               preds = preds,
                                               curr_formula=curr_formula)

                case _:
                    raise ValueError(f"*  _build_validation_excel:  Unknown validation rule {rule}")




    # RULE VALIDATION
    # ===============


        



    # EXCEL PRIMITIVES
    # ================

    # _add_values_row
    # Add a row of values to the model
    #  
    # INPUTS:
    #   
    #   tab = the tab to add it to
    #   values = the series of values to add
    #   restrictions = the set of restrictions on data entry
    # 
    # OUTPUTS:
    #   worksheet

    def _add_values_row(self, 
                        id: str,
                        tab_name: str,
                        values: list, 
                        curr_row_meta_data: ForecastMetaDataSeries, 
                        restriction: List[Dict[ForecastDataSeriesMetaDataValidationSchema, Any]] = None, 
                        row_name: str = None):
        
        # make sure it's a legit tab to go to start writing
        if(tab_name not in self.row_trackers.keys()):
            raise ValueError(f"* add_input_row:  error, requested tab '{tab_name}' not in the list of tabs\n{self.row_trackers.keys()}")
        
        # get meta_data values that we need for this function, if it wasn't provided
        if restriction is None:
            restriction = curr_row_meta_data.meta_data[ForecastMetaDataSeriesSchema.VALIDATION]
        
        # boilerplate set-up of the row
        (curr_cell, data_type, display_type) = self._add_row_setup(id, tab_name, curr_row_meta_data, row_name)

        # go cell by cell adding the row values
        for value in values:
            curr_cell.value = value # set the value of the cell
            self._build_validation_excel(curr_cell_meta_data = curr_row_meta_data, curr_cell = curr_cell, data_type = data_type, display_type = display_type, validation_rules = restriction) # add validation rules (if any)
            curr_cell = curr_cell.offset(row = 0, column = 1) # move the current cell pointer by one over to the right of the row


    
    # _add_arith_row
    # Add a row of arithmetic formulas (+, -, *, /)
    #  
    # INPUTS:
    #   
    #   tab = the tab to add it to
    #   values = the series of values to multiply
    #   type = the data type (for formatting purposes)
    #   restrictions = the set of restrictions on data entry
    # 
    # OUTPUTS:
    #   worksheet

    def _add_arith_row(self, 
                       arith_funct: ForecastBuilderExcelArithmeticFunctions, 
                       id: str, 
                       tab_name: str,
                       values: list,
                       curr_row_meta_data: ForecastMetaDataSeries,
                       preds: list[str] = None, 
                       restriction: List[Dict[ForecastDataSeriesMetaDataValidationSchema, Any]] = None, 
                       row_name: str = None):
        

        # make sure it's a legit tab to go to start writing
        if(tab_name not in self.row_trackers.keys()):
            raise ValueError(f"* add_prod_row:  error, requested tab '{tab_name}' not in the list of tabs\n{self.row_trackers.keys()}")
        
        # get meta_data values that we need for this function, if it wasn't provided
        if preds is None:
            preds = curr_row_meta_data.meta_data[ForecastMetaDataSeriesSchema.PRED]
        
        if restriction is None:
            restriction = curr_row_meta_data.meta_data[ForecastMetaDataSeriesSchema.VALIDATION]
        
        # boilerplate set-up of the row
        # and place the curr_cell pointer in the row and column to start adding values
        (curr_cell, data_type, display_type)  = self._add_row_setup(id, tab_name, curr_row_meta_data, row_name)

        # Iterate over all the columns in the row that need values
        num_vals_add = len(values)

        for curr_col in range(curr_cell.column, curr_cell.column + num_vals_add):
           list_of_formula_refs = self._ids_to_formula_refs(preds, col_num = curr_col)  # generate a list of formula referenes, and constants
           formula = f"={arith_funct.value.join(list_of_formula_refs)}"
           curr_cell.value = formula # create the formula (i.e. = 'Summary'!A1 + 'Summary'!B1 + 'Summary'!C1)
           
           # add validation rules (if any)
           self._build_validation_excel(curr_cell_meta_data = curr_row_meta_data, 
                                        curr_cell = curr_cell, 
                                        data_type = data_type, 
                                        display_type = display_type, 
                                        validation_rules = restriction,
                                        curr_formula = formula,
                                        apply_to_preds = True)
           
           curr_cell = curr_cell.offset(row = 0, column = 1)



    
    # _generate_core_tabs
    # Create the initial set of tabs required by the player
    #  
    # INPUTS:
    #   NA
    # 
    # OUTPUTS:
    #   NA

    def _generate_core_tabs(self):
        # get the number of worksheets in workbook
        self.template_num_sheets = len(self.player_model.sheetnames)
        
        # generate the core tabs of the model
        for tab in self.EXCEL_REQUIRED_WORKBOOK_TABS:
            ForecastExcelBaseHelpers.create_ws(tab, protect_worksheet = FORECAST_EXCEL_PROTECT_WORKSHEET, workbook = self.player_model, num_sheets = self.template_num_sheets)



    

    # VALIDATION & STYLES

    
    # handles setting any formatting and read/write access based on the input restrictions to this cell
    def _add_input_restrictions(self, 
                                curr_cell: Cell, 
                                data_type: ForecastDataSeriesMetaDataDataType, 
                                display_type: ForecastDataSeriesMetaDataDataType,
                                restriction: ForecastDataSeriesMetaDataValidateInputRestrictions,
                                curr_cell_meta_data: ForecastMetaDataSeries = None):

        # if token check is requested, the figure out if the current value matches the ForecastDataModel approved editable values token
        # if it does, make it read_write, otherwise make it read only
        if(restriction == ForecastDataSeriesMetaDataValidateInputRestrictions.TOKEN_CHECK):
            if(curr_cell.value == ForecastDataModel.EDITABLE_VALUES_TOKEN):
                restriction = ForecastDataSeriesMetaDataValidateInputRestrictions.READ_WRITE
            else:
                restriction = ForecastDataSeriesMetaDataValidateInputRestrictions.READ_ONLY

        # # based on the restriction and the data type, apply cell protection and the appropriate style
        ForecastExcelCellStyleBuilder.generate_cell_data_type_style(curr_cell = curr_cell, 
                                                                    display_type = display_type, 
                                                                    restriction = restriction,
                                                                    curr_cell_meta_data = curr_cell_meta_data)
        
        ForecastExcelValidationRuleBuilder.generate_data_entry_rule(curr_cell = curr_cell, 
                                                                    data_type = data_type,
                                                                    error_title = "Invalid entry",
                                                                    prompt = None,
                                                                    prompt_title = None,
                                                                    allow_blank = True,
                                                                    curr_cell_meta_data = curr_cell_meta_data)



    # handles a validation involving a comparison of an value to another value
    def _add_value_comparison(self,
                              curr_cell: Cell,
                              comparison_type: ForecastDataSeriesMetaDataComparisonType,
                              curr_cell_meta_data: ForecastMetaDataSeries,
                              preds : list[str] = None,
                              curr_formula: str = None):
            
            if(comparison_type not in [ForecastDataSeriesMetaDataComparisonType.BETWEEN, ForecastDataSeriesMetaDataComparisonType.NOT_BETWEEN]):
                # get the limit value for the comparison check
                comparison_id_or_const = curr_cell_meta_data.meta_data[ForecastMetaDataSeriesSchema.ARGS][comparison_type]

                ForecastExcelValidationRuleBuilder.generate_comparison_rule(curr_cell = curr_cell,
                                                                            comparison_type = comparison_type,
                                                                            value = comparison_id_or_const,
                                                                            preds = preds,
                                                                            error_title = "Invalid entry",
                                                                            prompt = None,
                                                                            prompt_title = None,
                                                                            allow_blank = True,
                                                                            curr_formula = curr_formula,
                                                                            curr_cell_meta_data = curr_cell_meta_data)
            else:
                # get the limit value for the comparison check
                min_comparison_id_or_const = curr_cell_meta_data.meta_data[ForecastMetaDataSeriesSchema.ARGS][comparison_type][0]
                max_comparison_id_or_const = curr_cell_meta_data.meta_data[ForecastMetaDataSeriesSchema.ARGS][comparison_type][1]

                ForecastExcelValidationRuleBuilder.generate_comparison_rule(curr_cell = curr_cell,
                                                            comparison_type = comparison_type,
                                                            value = min_comparison_id_or_const,
                                                            preds = preds,
                                                            error_title = "Invalid entry",
                                                            prompt = None,
                                                            prompt_title = None,
                                                            allow_blank = True,
                                                            value2 = max_comparison_id_or_const,
                                                            curr_formula = curr_formula,
                                                            curr_cell_meta_data = curr_cell_meta_data)




    

    # HELPER FUNCTIONS 
    # ================

    # _add_row_setup
    # Common boilerplate for all _add actions (values, arith function) to add a new row for an id
    # INPUT
    #   id - ForecastMetaDataSeries ID
    #   tab_name - Tab to add the row to
    #   curr_row_meta_data - ForecastMetaDataSeries information to grab additional row information
    #   row_name (optional) - if a row name (label) should be added, put it here
    #
    # OUTPUT:
    #   curr_cell - a reference to an excel spreadsheet pointing to the first location to write a value
    #   data_type - the data_type to be written
    #   display_type - the data_type to be displayed (for example, we may have a float, but needs to be displayed as an int)
    def _add_row_setup(self, id: str, tab_name: str, curr_row_meta_data: ForecastMetaDataSeries, row_name: str = None) -> Tuple[Any, ForecastDataSeriesMetaDataDataType, ForecastDataSeriesMetaDataDataType]:

        # grab meta-data that we always grab and return it so we don't have to do this in every function call
        data_type = curr_row_meta_data.meta_data[ForecastMetaDataSeriesSchema.DATA_TYPE]
        display_type = curr_row_meta_data.meta_data[ForecastMetaDataSeriesSchema.DISPLAY_TYPE]

        # set-up for the adding a row, get a worksheet for the tab name, and the next row to add to
        ws = self.player_model[tab_name]
        curr_row = self.row_trackers[tab_name]
        start_row_ref = ws.cell(row = curr_row, column = self.EXCEL_START_COL)
        
        # add any common row prefix stuff if needed:
        # row label
        if(row_name is not None):
            curr_cell = ws.cell(row = curr_row, column = self.EXCEL_LABEL_COL)
            self._add_label(value = row_name, curr_cell = curr_cell)

        # save the start cell ref for this id, and the tab_name (tab name not needed, but for convenience)
        self._track_add_row(id = id, tab_name = tab_name, cell_ref = start_row_ref) # store this 

        # move current cell to the start of the values entry
        curr_cell = ws.cell(row = curr_row, column = self.EXCEL_VALUES_START_COL)

        return(curr_cell, data_type, display_type)
   


    # _track_add_row
    # handles the boilerplate bookkeeping whenever a new row is added to the model
    #
    # INPUTS:
    #   id = ForecastMetaDataSeries id for the row
    #   tab_name = tab name for the row (mostly for convenience so we don't have to look it up in cell reference)
    #   cell_ref = (cell) cell object for the start of the row
    #
    # OUTPUTS:
    #

    def _track_add_row(self, id: str, tab_name: str, cell_ref: Cell):
        self.id_cellref_map.add(id = id, tab_name = tab_name, cell_ref = cell_ref)  # add the id of the row to our mapper
        self.row_trackers[tab_name] += 1  # increment our row tracker for this tab, so we know location of next row to add


    # _ids_to_formula_refs
    # given a list of either constants (int or float) or ForecastMetaData IDs (strings)
    # and a COLUMN in the spreadsheet, generates a string reference suitable for inclusion in a formula (i.e. 'Summary'!A1)
    # this is useful when you need to generate all the formula values in the same column, but different rows for a forumla
    # i.e. ('Summary'!A1 + 'Summary'!B1 + 'Summary'!C1)
    # NOTE:  the list HAS TO BE A COLUMN... because only one number is provided for the column offset, so if they aren't all lined up
    #        in the same column, the formula will be wrong.
    # 
    # INPUTS:
    #   list_of_ids - list with a mix of ForecastMetaDataSeries IDs or constants (ints or floats)
    #
    # OUTPUTS:
    #   list of strings - each string being either a constant (string version of float or int) or a cell reference suitable for an excel formula

    def _ids_to_formula_refs(self, list_of_ids: list[int | str | float], col_num: int, with_ws_name: bool = True) -> list[str]:
        list_of_refs = []

        # iterate of all the list
        for id_or_const in list_of_ids:
            list_of_refs.append(self._id_to_formula_ref(id_or_const, col_num, with_ws_name = with_ws_name))

        return(list_of_refs)
    

    # _id_to_formula_refs
    # given a value which is either constants (int or float) or ForecastMetaData IDs (strings)
    # and a column in the spreadsheet, generates a string reference suitable for inclusion in a formula (i.e. 'Summary'!A1)
    # this is useful when you need to generate all the formula values in the same column, but different rows for a forumla
    # i.e. ('Summary'!A1 + 'Summary'!B1 + 'Summary'!C1)
    # 
    # INPUTS:
    #   list_of_ids - list with a mix of ForecastMetaDataSeries IDs or constants (ints or floats)
    #   col_num = the ABSOLUTE column number for for the cell (we will then convert it to a relative number, which we will then use to offset the row location)
    #
    # OUTPUTS:
    #   list of strings - each string being either a constant (string version of float or int) or a cell reference suitable for an excel formula

    def _id_to_formula_ref(self, id_or_const: int | str | float, col_num: int, with_ws_name: bool = True) -> str:
            # determine if this is an ID or a constant
            # if constant, return a string version of the constant
            if isinstance(id_or_const, int | float):
                const = id_or_const
                return(str(const))

            # if ID, get the 
            else:
                id = id_or_const
                (tab_name, cell_ref) = self.id_cellref_map.get(id)  # get the tab_name and the cell reference
                col_offset = col_num - self.EXCEL_START_COL # figure out how many cells over we need to shift the reference to get our values
                cell_ref = cell_ref.offset(row = 0, column = col_offset)
                return(ForecastExcelBaseHelpers.cell_to_formula_ref(cell_ref, with_ws_name=with_ws_name))
            


    # LABELS
    def _add_label(self, value: str, curr_cell: Cell):
        curr_cell.value = value
        ForecastExcelCellStyleBuilder.generate_row_header_label(curr_cell)
        curr_cell.protection = Protection(locked = True, hidden = False)
