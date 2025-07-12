#####################################################################
# forecast_excel_base_helpers.py
#
#####################################################################


# OVERALL IMPORTS
# ===============

# FORECAST SPECIFIC IMPORTS
# =========================


# COMPONENT SPECIFIC IMPORTS
# ==========================
from openpyxl import Workbook, worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Protection
from langflow.base.forecasting_common.renderers.excel.forecast_excel_cell_style_builder import ForecastExcelCellStyleBuilder


class ForecastExcelBaseHelpers:
    # WORKSHEET

    # create_ws
    # Standardizes the creation of a tab, including adding headers and anything else need
    #  
    # INPUTS:
    #   tab = either the name of a worksheet or the 0-based index of the worksheet in a workbook
    # 
    # OUTPUTS:
    #   worksheet

    @staticmethod
    def create_ws(tab: str, protect_worksheet: bool, workbook: Workbook, num_sheets = int) -> worksheet:
        ws = ForecastExcelBaseHelpers.safe_create_ws(tab, workbook = workbook, num_sheets = num_sheets)
        ws['A1'] = tab
        ForecastExcelCellStyleBuilder.generate_ws_header(ws['A1'])

        if protect_worksheet:
            ForecastExcelBaseHelpers.protect_ws(ws)

        return(ws)



    # safe_create_ws
    # Simple convenience function to check if a worksheet already exists before creating it
    #  
    # INPUTS:
    #   tab = either the name of a worksheet or the 0-based index of the worksheet in a workbook
    # 
    # OUTPUTS:
    #   worksheet

    @staticmethod
    def safe_create_ws(tab: str, workbook: Workbook, num_sheets: int) -> worksheet:

        # check if a tab exists before creating it, if it does, leave it alone
        try:
            ws = workbook[tab]
        except:
            sheet_index = len(workbook.sheetnames) - num_sheets
            ws = workbook.create_sheet(tab, index = sheet_index)

        return(ws)



    # remove_ws
    # Simple convenience function to remove a sheet by either index or name using get_ws
    #  
    # INPUTS:
    #   tab = either the name of a worksheet or the 0-based index of the worksheet in a workbook
    # 
    # OUTPUTS:
    #   None

    @staticmethod
    def remove_ws(tab: str | int, workbook: Workbook):
        workbook.remove(ForecastExcelBaseHelpers.get_ws(tab))



    # get_ws
    # Simple convenience function to enable all other worksheet functions to access a worksheet by name or index
    #  
    # INPUTS:
    #   tab = either the name of a worksheet or the 0-based index of the worksheet in a workbook
    # 
    # OUTPUTS:
    #   worksheet

    @staticmethod
    def get_ws(tab: str | int, workbook: Workbook) -> worksheet:
        if isinstance(tab, str):
            ws = workbook[tab]
        elif isinstance(tab, int):
            ws = workbook[workbook.sheetnames[tab]]
        else:
            raise ValueError(f"*  get_ws:  invalid 'tab' type: {type(tab)}, value: {tab}")
        
        return(ws)


    # protect_ws
    # Protect a worksheet
    #  
    # INPUTS:
    #   ws - Worksheet object
    # 
    # OUTPUTS:
    #   NA

    @staticmethod
    def protect_ws(ws: worksheet):
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        ws.protection.sheet = True




    # CELL

    # cell_to_formula_ref
    # Given a cell object, return a fully qualified string (i.e. with tab name) suitable for inclusion in a formula (i.e. 'Summary'!A1)
    # 
    # INPUTS:
    #   cell reference object
    #   with_ws_name - (optional) if True, include the worksheet name in the reference, otherwise just return the cell coordinate
    #
    # OUTPUTS:
    #   string - a cell reference suitable for an excel formula (i.e. 'Summary'!A1)
    
    @staticmethod
    def cell_to_formula_ref(cell: Cell, with_ws_name = True) -> str:
        if(with_ws_name):
            return(f"'{cell.parent.title}'!{cell.coordinate}")
        else:
            return(cell.coordinate)
        

    # FORMULA MANIPULATION

    # find_comma_with_the_balanced_parens
    # In excel formulas that have the form
    # (sub_term1), (sub_term2), (sub_term3)...
    # where each sub_term can also have parantheses and commas, split out only the highest level sub_terms and return a list
    # INPUTS
    #  text = text to parse
    #
    # OUTPUT
    #   list[str] = list of each sub_term
    @staticmethod
    def find_comma_with_balanced_parens(text: str) -> list[str]:
        results = []
        paren_balance = 0
        start_index = 0
        for i, char in enumerate(text):
            if char == '(':
                paren_balance += 1
            elif char == ')':
                paren_balance -= 1
            elif char == ',' and paren_balance == 0:
                results.append(text[start_index:i])
                start_index = i + 1
        # Add the last part after the last comma or if no comma exists
        if start_index < len(text):
            results.append(text[start_index:])
        return results
    


    # remove_worksheet_names_from_formula
    # In excel formula convert all cell references that HAVE a worksheet name in the reference to ones which do one
    # i.e. ='Summary'!A1 + 'Summary'!A2 -> =A1 + A2
    # INPUTS
    #   formula = excel formula as a string to parse
    #
    # OUTPUT
    #   formula string with cell references without worksheet names
    @staticmethod
    def remove_worksheet_names_from_formula(formula: str) -> str:
        """
        Removes worksheet names from cell references in an Excel formula string.
        Example: "='Summary'!A1 + 'Sheet2'!B2 + C3" -> "A1 + B2 + C3"
        """
        import re
        # Regex matches: optional single-quoted or unquoted worksheet name followed by '!' and a cell reference
        # Handles cases like 'Summary'!A1, Sheet2!B2, etc.
        return re.sub(r"(?:'[^']+'|[A-Za-z0-9_]+)!", "", formula)
    


    # convert_formula_to_sub_term
    # Takes a formula, removes the "=" at the from, and surrounds with parens, making it
    # suitable for inclusion in another formula
    # INPUTS
    #   formula = excel formula as a string to parse
    #
    # OUTPUT
    #   formula sub term
    @staticmethod
    def convert_formula_to_sub_term(formula: str) -> str:
        if(formula[0] == "="):
            formula = formula[1:]
        
        return(f"({formula})")

