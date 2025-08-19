#####################################################################
# forecast_excel_base_helpers.py
#
#####################################################################


# OVERALL IMPORTS
# ===============

# FORECAST SPECIFIC IMPORTS
# =========================


# COMPONENT SPECIFIC IMPORTS
# ==========================
from openpyxl import Workbook, worksheet
from openpyxl.cell.cell import Cell
from openpyxl.styles import Protection
from langflow.base.forecasting_common.builders.excel.forecast_excel_cell_style_builder import ForecastExcelCellStyleBuilder


class ForecastExcelBaseHelpers:
    # WORKSHEET

    # create_ws
    # Standardizes the creation of a tab, including adding headers and anything else need
    #  
    # INPUTS:
    #   tab = name of the worksheet
    #   protect_worksheet = should all the cells in the worksheet be protected
    #   workbook = the openpyxl Workbook object which will hold this worksheet
    #   num_sheets = (optional) how many sheets FROM THE END to insert this workbook
    #   
    # 
    # OUTPUTS:
    #   worksheet

    @staticmethod
    def create_ws(tab: str, protect_worksheet: bool, workbook: Workbook, num_sheets: int = 0) -> worksheet:
        ws = ForecastExcelBaseHelpers.safe_create_ws(tab, workbook = workbook, num_sheets = num_sheets)
        ws['A1'] = tab
        ForecastExcelCellStyleBuilder.generate_ws_header(ws['A1'])

        if protect_worksheet:
            ForecastExcelBaseHelpers.protect_ws(ws)

        return(ws)



    # safe_create_ws
    # Simple convenience function to check if a worksheet already exists before creating it
    #  
    # INPUTS:
    #   tab = either the name of a worksheet or the 0-based index of the worksheet in a workbook
    # 
    # OUTPUTS:
    #   worksheet

    @staticmethod
    def safe_create_ws(tab: str, workbook: Workbook, num_sheets: int) -> worksheet:

        # check if a tab exists before creating it, if it does, leave it alone
        try:
            ws = workbook[tab]
        except:
            sheet_index = len(workbook.sheetnames) - num_sheets
            ws = workbook.create_sheet(tab, index = sheet_index)

        return(ws)



    # remove_ws
    # Simple convenience function to remove a sheet by either index or name using get_ws
    #  
    # INPUTS:
    #   tab = either the name of a worksheet or the 0-based index of the worksheet in a workbook
    # 
    # OUTPUTS:
    #   None

    @staticmethod
    def remove_ws(tab: str | int, workbook: Workbook):
        workbook.remove(ForecastExcelBaseHelpers.get_ws(tab))



    # get_ws
    # Simple convenience function to enable all other worksheet functions to access a worksheet by name or index
    #  
    # INPUTS:
    #   tab = either the name of a worksheet or the 0-based index of the worksheet in a workbook
    # 
    # OUTPUTS:
    #   worksheet

    @staticmethod
    def get_ws(tab: str | int, workbook: Workbook) -> worksheet:
        if isinstance(tab, str):
            ws = workbook[tab]
        elif isinstance(tab, int):
            ws = workbook[workbook.sheetnames[tab]]
        else:
            raise ValueError(f"*  get_ws:  invalid 'tab' type: {type(tab)}, value: {tab}")
        
        return(ws)


    # protect_ws
    # Protect a worksheet
    #  
    # INPUTS:
    #   ws - Worksheet object
    # 
    # OUTPUTS:
    #   NA

    @staticmethod
    def protect_ws(ws: worksheet):
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        ws.protection.sheet = True




    # gen_excel_tab_name
    @staticmethod
    def gen_excel_tab_name(name: str, existing_tab_names: list[str]) -> str:
        MAX_TAB_ID_NUM = 1000
        num_tries = 0

        # generate a candidate tab name
        tab_name = ForecastExcelBaseHelpers._gen_excel_tab_name_candidate(name = name)

        # if candiate name already exists and a counter to the name (i.e."name1", "name2", ...) and keep iterating till we find
        # one that works (or we have tried 999 names)
        while(tab_name in existing_tab_names):
            num_tries += 1

            if(num_tries >= MAX_TAB_ID_NUM):
                raise ValueError(f"\n*  gen_excel_tab_name:  unable to generate unique tab name for {name}, no unique name up to '999'")

            tab_name = ForecastExcelBaseHelpers._gen_excel_tab_name_candidate(name = name, id = num_tries)
        
        return tab_name
            



    # gen_excel_tab_name_candidate
    # takes any variable and generates a name suitable which will fit in an excel tab as a name
    # this is not guaranteed to be unique, but if given a last_id for the name, will increment the last id
    # by one
    @staticmethod
    def _gen_excel_tab_name_candidate(name: str, id: int = None) -> str:
        chars_to_remove = 0

        # if we have an id, figure out how many character it will need in the tab name
        # (tabs in EXCEL are limited to 31 chars, so any id characters must decrement from that 31)
        if(id is not None):
            chars_to_remove = len(str(id))

        # Replace invalid characters
        invalid_chars = r'\/?:*[]'
        for char in invalid_chars:
            name = name.replace(char, '_')  # Replace with underscore or other suitable character

        # Remove leading/trailing apostrophes if present
        if name.startswith("'"):
            name = name[1:]
        if name.endswith("'"):
            name = name[:-1]

        # Truncate if longer than 31 characters
        if len(name) > 31:
            if id is None:
                name = name[:31]
            else:
                name = name[:(31-chars_to_remove)]    # remove enough characters to be able to add the id

        # Ensure the name is not empty after cleaning
        if not name:
            name = "Sheet"  # Default name if cleaning results in an empty string

        if id is not None:
            name = f"{name}{id}"

        return name
    

    # CELL

    # cell_to_formula_ref
    # Given a cell object, return a fully qualified string (i.e. with tab name) suitable for inclusion in a formula (i.e. 'Summary'!A1)
    # 
    # INPUTS:
    #   cell reference object
    #   with_ws_name - (optional) if True, include the worksheet name in the reference, otherwise just return the cell coordinate
    #
    # OUTPUTS:
    #   string - a cell reference suitable for an excel formula (i.e. 'Summary'!A1)
    
    @staticmethod
    def cell_to_formula_ref(cell: Cell, with_ws_name = True) -> str:
        if(with_ws_name):
            return(f"'{cell.parent.title}'!{cell.coordinate}")
        else:
            return(cell.coordinate)
        

    # FORMULA MANIPULATION

    # find_comma_with_the_balanced_parens
    # In excel formulas that have the form
    # (sub_term1), (sub_term2), (sub_term3)...
    # where each sub_term can also have parantheses and commas, split out only the highest level sub_terms and return a list
    # INPUTS
    #  text = text to parse
    #
    # OUTPUT
    #   list[str] = list of each sub_term
    @staticmethod
    def find_comma_with_balanced_parens(text: str) -> list[str]:
        results = []
        paren_balance = 0
        start_index = 0
        for i, char in enumerate(text):
            if char == '(':
                paren_balance += 1
            elif char == ')':
                paren_balance -= 1
            elif char == ',' and paren_balance == 0:
                results.append(text[start_index:i])
                start_index = i + 1
        # Add the last part after the last comma or if no comma exists
        if start_index < len(text):
            results.append(text[start_index:])
        return results
    


    # remove_worksheet_names_from_formula
    # In excel formula convert all cell references that HAVE a worksheet name in the reference to ones which do one
    # i.e. ='Summary'!A1 + 'Summary'!A2 -> =A1 + A2
    # INPUTS
    #   formula = excel formula as a string to parse
    #
    # OUTPUT
    #   formula string with cell references without worksheet names
    @staticmethod
    def remove_worksheet_names_from_formula(formula: str) -> str:
        """
        Removes worksheet names from cell references in an Excel formula string.
        Example: "='Summary'!A1 + 'Sheet2'!B2 + C3" -> "A1 + B2 + C3"
        """
        import re
        # Regex matches: optional single-quoted or unquoted worksheet name followed by '!' and a cell reference
        # Handles cases like 'Summary'!A1, Sheet2!B2, etc.
        return re.sub(r"(?:'[^']+'|[A-Za-z0-9_]+)!", "", formula)
    

    # remove_worksheet_names_from_string
    # Alias for 'remove_worksheet_names_from_formula'
    @staticmethod
    def remove_worksheet_names_from_string(ref: str) -> str:
        return(ForecastExcelBaseHelpers.remove_worksheet_names_from_formula(ref))  
    


    # remove_worksheet_names_from_list
    # Given a list of strings, all of which are excel cell reference ready for formulas (i.e. 'Summary'!A1), remove the worksheet name in the reference
    # INPUTS
    #   formula = excel formula as a string to parse
    #
    # OUTPUT
    #   formula string with cell references without worksheet names
    @staticmethod
    def remove_worksheet_names_from_list(list_of_refs: list[str]) -> list[str]:
        list_of_refs_without_worksheet_names = []

        for full_ref_str in list_of_refs:
            list_of_refs_without_worksheet_names.append(ForecastExcelBaseHelpers.remove_worksheet_names_from_string(full_ref_str))

        return(list_of_refs_without_worksheet_names)
    
    
    # convert_formula_to_sub_term
    # Takes a formula, removes the "=" at the from, and surrounds with parens, making it
    # suitable for inclusion in another formula
    # INPUTS
    #   formula = excel formula as a string to parse
    #
    # OUTPUT
    #   formula sub term
    @staticmethod
    def convert_formula_to_sub_term(formula: str) -> str:
        if(formula[0] == "="):
            formula = formula[1:]
        
        return(f"({formula})")
    


    # META-DATA

    # quick_static_date_series
    # generates a date ForecastMetaDataSeries with the minimum of inputs, used by the builder to quickly generate inputs for "action_DATES" calls
    @staticmethod
    def quick_static_date_series(step, label, values, id: str = None, key_length: int = 5):
        from langflow.base.forecasting_common.models.forecast_meta_data import ForecastMetaDataSeriesIdGenerator, ForecastMetaDataSeries, ForecastDataSeriesMetaDataAction, ForecastDataSeriesMetaDataDataType, ForecastDataSeriesMetaDataValidationSchema, ForecastDataSeriesMetaDataValidateInputRestrictions
        
        if(id is None):
            id = ForecastMetaDataSeriesIdGenerator.static_gen_rel_id(prefix = f"quick_static_date_series_{ForecastMetaDataSeries}", length = key_length)

        return(
            ForecastMetaDataSeries(id = id,
                                   step_type = step,
                                   action = ForecastDataSeriesMetaDataAction.DATES,
                                   data_type = ForecastDataSeriesMetaDataDataType.DATE,
                                   display_type = ForecastDataSeriesMetaDataDataType.DATE,
                                   display_name = label,
                                   data_values = values,
                                   validation = [{ForecastDataSeriesMetaDataValidationSchema.INPUT_RESTRICTION: ForecastDataSeriesMetaDataValidateInputRestrictions.READ_ONLY}],)
        )
    

    @staticmethod
    def quick_static_input_series(step, label, values, id: str = None, key_length: int = 5):
        from langflow.base.forecasting_common.models.forecast_meta_data import ForecastMetaDataSeriesIdGenerator, ForecastMetaDataSeries, ForecastDataSeriesMetaDataAction, ForecastDataSeriesMetaDataDataType, ForecastDataSeriesMetaDataValidationSchema, ForecastDataSeriesMetaDataValidateInputRestrictions

        if(id is None):
            id = ForecastMetaDataSeriesIdGenerator.static_gen_rel_id(prefix = f"quick_static_input_series_{ForecastMetaDataSeries}", length = key_length)

        return(
            ForecastMetaDataSeries(id = id,
                                   step_type = step,
                                   action = ForecastDataSeriesMetaDataAction.INPUT,
                                   data_type = ForecastDataSeriesMetaDataDataType.INT,
                                   display_type = ForecastDataSeriesMetaDataDataType.INT,
                                   display_name = label,
                                   data_values = values,
                                   validation = [{ForecastDataSeriesMetaDataValidationSchema.INPUT_RESTRICTION: ForecastDataSeriesMetaDataValidateInputRestrictions.TOKEN_CHECK}],)
        )
    
