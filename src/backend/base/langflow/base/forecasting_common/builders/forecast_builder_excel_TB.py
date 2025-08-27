#####################################################################
# forecast_builder_excel_TB.py
#
# Implements the a summation component.  It's already implemented everywhere
# this just makes it explicit (for visual presentation purposes)
# 
# INPUTS:  DataFrame
# OUTPUTS:  DataFrame
#
#####################################################################


from typing import List, Dict, Tuple, Any
from datetime import datetime
import pandas as pd
import numpy as np
from langflow.schema.dataframe import DataFrame, Data


# FORECAST SPECIFIC IMPORTS
# =========================
from langflow.base.forecasting_common.constants import FORECAST_INT_TO_SHORT_MONTH_NAME, ForecastModelInputTypes, ForecastModelTimescale

from langflow.base.forecasting_common.models.forecast_data_model import ForecastDataModel
from langflow.base.forecasting_common.models.forecast_meta_data import (ForecastMetaDataSeries, 
                                                                        ForecastMetaDataFrame, 
                                                                        ForecastMetaDataSeriesSchema, 
                                                                        ForecastMetaDataFrameSchema,
                                                                        ForecastDataSeriesMetaDataStepTypes, 
                                                                        ForecastDataSeriesMetaDataAction, 
                                                                        ForecastDataSeriesMetaDataDataType, 
                                                                        ForecastDataSeriesMetaDataValidationSchema, 
                                                                        ForecastDataSeriesMetaDataValidateInputRestrictions,
                                                                        ForecastDataSeriesMetaDataComparisonType,
                                                                        ForecastDataSeriesMetaDataArgsTreatmentStepInit,
                                                                        ForecastDataSeriesMetaDataArgsMTYYTM)

from langflow.base.forecasting_common.models.forecast_data_interface import IdElementToHandleMap, IdElementToHandleMaps, ForecastPredRef, ForecastPredIterator



# COMPONENT SPECIFIC IMPORTS
# ==========================
from datetime import datetime
from enum import Enum
import shutil
from openpyxl import Workbook, worksheet, load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection

from  langflow.components.forecasting_TB.forecast_epidemiology_TB import FORECAST_EPIDEMIOLOGY_DATES_LABEL

from langflow.base.forecasting_common.builders.excel.forecast_excel_base_helpers import ForecastExcelBaseHelpers
from langflow.base.forecasting_common.builders.excel.forecast_excel_validation_builder import ForecastExcelValidationRuleBuilder
from langflow.base.forecasting_common.builders.excel.forecast_excel_cell_style_builder import ForecastExcelCellStyleBuilder

from langflow.components.forecasting_TB.forecast_treatment_TB import ForecastTreatmentStepInitArgs



# CONFIG STUFF
FORECAST_EXCEL_PROTECT_WORKSHEET = False


# CLASSES
# =======


# Enum of EXCEL_ARITHMETIC_FUNCTIONS
class ForecastBuilderExcelArithmeticFunctions(str, Enum):
    ADD = " + "
    SUB = " - "
    PROD = " * "
    DIV = " / "

# Enum of EXCEL Required Tabs
class ForecastBuilderExcelRequiredTabs(str, Enum):
    #INPUTS = "Inputs"
    SUMMARY = "Summary"
    #TIME_CONVERT = "Time Convert"




# IdToCellReferenceMap
# map ForecastMetaDataFrame id's to the cell locations that they represent
class IdToCellReferenceMap(IdElementToHandleMap):

    # __init__ function, does nothing right now
    def __init__(self, default_num_elements: int):

        # INSTANCE VARIABLES
        self.default_num_elements = default_num_elements
        self.id_to_ref_map = {}

    # Add a new entry to the map: id is the key, tab_name and cell references are the values
    # TODO:  change the returned value of 'Any' into the name of the 'cell' object
    def add(self, id: str, tab_name: str, cell_ref: Cell, num_elements: int = None):
        if(id in self.id_to_ref_map.keys()):
            raise ValueError(f"\n* IdToCellReferenceMap.add: error, id {id} already exists in map:\n{self.id_to_ref_map.keys()}")

        if num_elements is None:
            num_elements = self.default_num_elements

        self.id_to_ref_map[id] = {"tab": tab_name, "ref": cell_ref, "num_elements": num_elements}


    # Given an id, return the: tab_name, cell object pointing to the start of row
    # TODO:  change the returned value of 'Any' into the name of the 'cell' object
    def get(self, id: str) -> Tuple[str, Any, int]:
        try:
            full_ref = self.id_to_ref_map[id]
            tab_name = full_ref["tab"]
            cell_ref = full_ref["ref"]
            num_elements = full_ref["num_elements"]
        except:
            raise ValueError(f"\n* IdToCellReferenceMap.get: error, id {id} not found in map:\n{self.id_to_ref_map.keys()}")
        
        return(tab_name, cell_ref, num_elements)
    

    # Return a PRED ref, which is more complicated than a simple get
    def ref_to_obj(self, id: ForecastPredRef, element_num: int) -> tuple[str, Any, int]:
        if id.const is not None:
            raise ValueError(f"\n*  ref_to_obj:  invalid ForecastPrefRef passed, ref is a constant {id.const}.")
        
        # get the reference
        (tab_name, cell_ref, num_elements) = self.get(id.rel_id)

        # figure out which element number of return
        if(id.has_single_value):
            element_idx = id.single_value
        else:
            element_idx = element_num

        if(id.has_shift_value):
            element_idx += id.shift_value

        # handle the element being outside of bounds
        if (element_idx > num_elements) or (element_idx < 0):
            # for convenience, it is expected when using shift that elements may fall outside of bounds,
            # so if that is the case, don't raise an error, just return zero
            if(id.has_shift_value):
                return(tab_name, 0, num_elements)
            else:
                raise ValueError(f"\n*  ref_to_obj:  element number requested {element_num} is outside of id ({id.full_id}_{id.rel_id}) bounds ({num_elements}).")
            
        # if element inside bounds, return the object
        else:
            if(element_idx > 0):
                cell_ref= cell_ref.offset(row = 0, column = element_idx)

            # Add the VALUE offset as well (since PREDs are always values) ZIV
            cell_ref = cell_ref.offset(row = 0, column = (ForecastBuilderExcelTB.ExcelField.VALUES - ForecastBuilderExcelTB.ExcelField.START))

            return(tab_name, cell_ref, num_elements)
        
    
    # get a list of ids in the map
    def get_ids(self) -> list[str]:
        return list(self.id_to_ref_map.keys())





# map ForecastMetaDataFrame id's to the cell locations that they represent
class IdToCellReferenceMaps(IdElementToHandleMaps):

    # # CLASS VARIABLES
    # ref_maps = None

    def __init__(self, default_num_elements: int, default_ref_map_id: str = None):
        # INSTANCE VARIABLES
        self.ref_maps = {}
        self.default_ref_map_id = None
        self.player_model = None

        if(default_ref_map_id is not None):
            self.create_ref_map(ref_map_id = default_ref_map_id, default_num_elements = default_num_elements, is_default = True)



    def add(self, id: str, tab_name: str, cell_ref: Cell, ref_map_id: str = None, num_elements: int = None):
        if ref_map_id is None:
            if(self.default_ref_map_id is None):
                raise ValueError(f"\n* add: error, method called without ref_map_id, and no default ref_map_id available {id} {tab_name} {cell_ref}.")
            else:
                return self.ref_maps[self.default_ref_map_id].add(id, tab_name, cell_ref, num_elements)
        else:
            if ref_map_id not in self.ref_maps.keys():
                self.create_ref_map(ref_map_id = ref_map_id, default_num_elements = num_elements if num_elements is not None else 0)
                
            return self.ref_maps[ref_map_id].add(id, tab_name, cell_ref, num_elements)



    def get(self, id: str, ref_map_id: str = None) -> Tuple[str, Cell, int]:
        (ref_map_id, cell_id, cell_offset) = self._parse_id(id)

        if ref_map_id in self.ref_maps.keys():
            (tab_name, cell_ref, num_elements) = self.ref_maps[ref_map_id].get(cell_id)
        else:
            raise ValueError(f"\n* get: error, invalid ref_map '{ref_map_id}' provided in {id}, current ref maps {self.ref_maps.keys()}.")

        # if there is a cell offset, need to provide a cell address with is offset (to the right) by the offset amount
        if cell_offset > 0:
            cell_ref = cell_ref.offset(column = cell_offset)

        return(tab_name, cell_ref, num_elements)


        
    def create_ref_map(self, ref_map_id: str, default_num_elements: int, is_default = False):
        if(ref_map_id in self.ref_maps.keys()):
            raise ValueError(f"\n* add_ref_map:  error, attempting to create duplicate ref_map {ref_map_id}.")
        else:
            self.ref_maps[ref_map_id] = IdToCellReferenceMap(default_num_elements = default_num_elements)
            
            if is_default:
                self.default_ref_map_id = ref_map_id


    def get_all_ids(self) -> list[str]:
        all_ids = []

        for ref_map_id in self.ref_maps.keys():
            all_ids.extend([f"{ref_map_id}.{element_id}" for element_id in self.ref_maps[ref_map_id].get_ids()])

        return all_ids
    

    def get_map_ids(self) -> list[str]:
        return list(self.ref_maps.keys())
    
    
    def get_map_ids(self, ref_map_id: str = None) -> list[str]:
        if ref_map_id is None:
            if(self.default_ref_map_id is None):
                raise ValueError(f"\n* get_map_ids: error, method called without ref_map_id, and no default ref_map_id available.")
            else:
                return list(self.ref_maps[self.default_ref_map_id].get_ids())
        else:
            if ref_map_id in self.ref_maps.keys():
                return list(self.ref_maps[ref_map_id].get_ids())
            else:
                raise ValueError(f"\n* get_map_ids: error, invalid ref_map_id '{ref_map_id}' provided, current ref maps {self.ref_maps.keys()}.")



    

    # take an ForecastPrefRef and the element in the range current at and return a Cell object
    # INPUT
    #   id - id of the row requested
    #   element_num - the number of the element in that row being requested (0-index... I think)
    #
    # OUTPUT
    #   tab_name - (str) tab name in excel of the tab holding this cell
    #   cell - (Cell) the openpyxl Cell class holding the reference to the cell
    #   num_elements - (int) total number of elements in the row requested    tab_name, cell_ref, num_elements
    #   
    def ref_to_obj(self, id: ForecastPredRef, element_num: int) -> tuple[str, Any, int]:
        return self.ref_maps[id.full_id].ref_to_obj(id, element_num)



    # TODO:  ZIV       
    # should probably be moved to forecast_meta_data
    def _parse_id(self, id: str) -> tuple[str, str, int | None]:
        raise ValueError(f"\n*  _parse_id:  error, this function should not be called, it is not implemented in IdToCellReferenceMaps.")
    
        ref_map_id = None
        cell_id = None
        cell_offset = None

        # GET THE REF_MAP ID
        results_ref = id.split('.')

        if(len(results_ref) == 2):
            ref_map_id = results_ref[0]
            cell_part = results_ref[1]

        elif(len(results_ref) == 1):
            if(self.default_ref_map_id == ""):
                raise ValueError(f"\n* _parse_id: error, id provided without ref_map_id, and no default ref_map_id available {id}.")
            else:
                ref_map_id = self.default_ref_map_id
            
            cell_part = results_ref[0]

        else:   # results_ref > 2
            raise ValueError(f"\n*  _parse_id:  error, improperly formatted id provided ({id}), too many '.'.")
        
        # GET THE ROW OFFSET NUMBER
        results_offset = cell_part.split(":")

        if(len(results_offset) == 2):
            cell_id = results_offset[0]
            cell_offset = int(results_offset[1])

        elif(len(results_offset) == 1):
            cell_id = results_offset[0]
            cell_offset = 0

        else:   # results_offset > 2
            raise ValueError(f"\n*  _parse_id:  error, improperly formatted id provided ({id}), too many '.'.")

        return(ref_map_id, cell_id, cell_offset)
    








    
        

# ForecastBuilderExcelTB
# Class which builds a Forecast Model player for excel using an Time Based model
class ForecastBuilderExcelTB():
    # CONSTANTS
    # =========

    # FORECAST DATA MODEL CONSTANTS
    DATAMODEL_ACTION_COL = "action"
    DATAMODEL_PRED_COL = "input_rows"


    # PLAYER REQUIRED TABS
    NEW_TAB_INSERT_INDEX_FROM_END = 1   # the place (from the end of all the tabs to insert a new tab, this allows us to put new tabs AHEAD of any config tabs we may have)
    DEFAULT_CARD = ForecastBuilderExcelRequiredTabs.SUMMARY

    # PLAYER ROW LAYOUT  NOTE:  row numbers are 1-index like excel, not 0-index like python (per openpyxl)
    EXCEL_START_ROW = 4 # The first row of a worsheet for any building,
    
    # COLUMN WIDTH CONSTANTS (these constants are set based on an average char width of 7 pixels per character)
    AVG_CHAR_WIDTH_IN_PIXELS = 7 # 3360/480
    LABEL_COL_WIDTH = 480 / AVG_CHAR_WIDTH_IN_PIXELS # 480 pixels = 70 chars long


    # THE COLUMN LAYOUT FOR EXCEL ROW, NOTE:  column numbers are 1-index like excel, not 0-index like python (per openpyxl)
    class ExcelField(int, Enum):
        START = 2                       # The first row of a worsheet for any building
        LABEL = START                   # Row label
        ID = LABEL + 1                  # ForecastMetaDataSeries ID for row
        NAME = ID + 1                   # User ID who is entered the data into row
        COUNTRY = NAME + 1              # Country
        PRODUCT = COUNTRY + 1           # Product
        INDICATION = PRODUCT + 1        # Indication for product (if needed)
        VALUES = INDICATION + 1         #  The first column where forecast values start

    EXCEL_FIELD_TO_LETTER_MAP = {}      # Mapping table from ExcelField column numbers to Excel column letters/cell-refs (set-up during _initialize_new_builder)


    # VARIABLES
    # =========

    # CLASS VARIABLES
    # ---------------



    # INSTANCE VARIABLES
    # ------------------
    # start_year - start year of forecast
    # start_month - start of fiscal year (or 1 - January for calendar year)
    # timescale - the minimum unit for a period in the forecast (can be MONTH or YEAR)
    # num_periods - the number of periods in the forecast
    # forecast_model - the forecast model generated by the DESIGNER
    # output_location - the location to put the player
    # hasTemplate - True if the builder uses a template, False if not
    # template_location (optional) - the location to load the template
    # template (optional) - a template to use when developing the player
    # template_num_sheets - the number of sheets that came with the template (all our sheets must go in front of those)

    # player_model - the object model for the player being developed (openpyxl.Workbook class)
    # row_trackers - dict of ints, each tab in the spreadsheet is the key, and the current location to add a new row is the int
    # id_cellref_maps - holds the all the ref_maps which map from all IDs to excel cell references (full references including tab name and cell coordinates)



    # CONSTRUCTOR
    # ===========
    # The constructor gets all the variables which are specific to this particular builder vs generic to all implementers of RENDER INTERFACE.
    # In the case of this excel builder, this include if we are using a template and where that template is located, as well as the location to
    # save the output file to
    #
    # INPUTS:
    #   output_location - location to save the output player
    #   template_location (optional) - if we are using a template file, where is the file located
    # 
    # OUTPUTS:
    #   N/A

    def __init__(self,
                 data_frame: DataFrame | pd.DataFrame,
                 meta_data:  ForecastMetaDataFrame,
                 output_location: str,
                 template_location: str = None):

        self.data_frame = data_frame
        self.meta_data = meta_data
        
        self.output_location = output_location

        if(template_location is None):
            self.hasTemplate = False
            self.template_location = None
        else:
            self.hasTemplate = True
            self.template_location = template_location

        



    
    # BUILD INTERFACE FUNCTIONS
    # =========================

    # build_player
    # BUILD INTERFACE:  build a player and return to the caller function 
    #  
    # INPUTS:
    #   TBD
    # 
    # OUTPUTS:
    #   TBD
    def build_player(self):

        # save these as instance variables
        self.id = self.meta_data.get_id()
        self.start_year = self.meta_data.get_start_year()
        self.start_month = self.meta_data.get_start_month()
        self.num_periods = self.meta_data.get_num_periods()
        self.timescale = self.meta_data.get_timescale()
        self.input_type = self.meta_data.get_input_type()

        # initial setup when building
        self._initialize_new_builder(default_num_elements = self.num_periods)

        # generate the model
        self._build_model_excel()

        # save the model
        self._finalize_new_builder()



    # _build_metadataframe_model
    # The main dispathcher for the builder, iterates over all the rows in the ForecastMetaDataFrame model, and dispatches
    # different functions the handle the different ACTIONS to build the excel player
    #  
    # INPUTS:
    #   meta_data - (ForecastMetaDataFrame) the Frame whose series to iterate over to build the model
    #   default_card - (str) In excel, the tab to render all rows to
    #   return_final_id_to_default - (bool) In excel, take the final row_id and put it in the overall default_card for the player
    # 
    # OUTPUTS:
    #   NA
    def _build_metadataframe_model(self, meta_data: ForecastMetaDataFrame, default_card: str):
        # get the object id so that all local_refs will be stored in the correct map
        def_ref_map_id = meta_data.get_id()

        # iterate through the forecast model, column by column, dispatching as needed
        for curr_col in meta_data.values():
            # get the next ForecastMetaDataSeries in the model
            id = curr_col.get_id() # get the current ForecastMetaDataSeries

            # dispatch based on the action
            match curr_col.get_action():
                case ForecastDataSeriesMetaDataAction.VALUES:
                    self.action_VALUES(id, default_card, curr_col, ref_map_id = def_ref_map_id)

                case ForecastDataSeriesMetaDataAction.DATES:
                    self.action_DATES(id, default_card, curr_col, ref_map_id = def_ref_map_id)

                case ForecastDataSeriesMetaDataAction.INPUT:
                    self.action_INPUT(id, default_card, curr_col, ref_map_id = def_ref_map_id)

                # case ForecastDataSeriesMetaDataAction.COPY:             # TODO:  implement copy action (DO WE REALLY NEED THIS?)
                #     print(f"SHIFT not implemented: {id}")
                #     break

                case ForecastDataSeriesMetaDataAction.SUM:
                    self.action_SUM(id, default_card, curr_col, ref_map_id = def_ref_map_id)

                case ForecastDataSeriesMetaDataAction.TOTAL:
                    self.action_TOTAL(id, default_card, curr_col, ref_map_id = def_ref_map_id)

                case ForecastDataSeriesMetaDataAction.PROD:
                    self.action_PROD(id, default_card, curr_col, ref_map_id = def_ref_map_id)

                case ForecastDataSeriesMetaDataAction.SUB:
                    self.action_SUB(id, default_card, curr_col, ref_map_id = def_ref_map_id)

                # case ForecastDataSeriesMetaDataAction.SHIFT:            # TODO:  implement shift action
                #     print(f"SHIFT not implemented: {id}")
                #     break

                case ForecastDataSeriesMetaDataAction.STEP_INIT:
                    new_card = self.action_STEP_INIT(id, default_card, curr_col, ref_map_id = def_ref_map_id)

                    # if we are switching cards, put the last_previous line in the SUMMARY tab (the self.default_card tab)
                    if(new_card != default_card):
                        # TODO
                        pass

                    #set the default_card to the new card 
                    default_card = new_card


                case ForecastDataSeriesMetaDataAction.YEAR_TO_MONTH:    # TODO:  implement year_to_month action
                    self.action_YEAR_TO_MONTH(id, default_card, curr_col, ref_map_id = def_ref_map_id, add_blank_row_before = True)

                case ForecastDataSeriesMetaDataAction.MONTH_TO_YEAR:    # TODO:  implement month_to_year action
                    self.action_MONTH_TO_YEAR(id, default_card, curr_col, ref_map_id = def_ref_map_id, add_blank_row_before = True)

                case _:
                    raise ValueError(f"\n*  _build_metadataframe_model:  Unknown action {curr_col.get_action()}")
        



    # action_VALUES
    # BUILD INTERFACE:  handle the VALUES action by creating a row of whatever values are provided
    #  
    # INPUTS:
    #   id = the id of the row
    #   card_name = name of the EXCEL tab to output results
    #   curr_meta_col = ForecastMetaDataSeries meta_data for this column
    # 
    # OUTPUTS:
    #   NA
    def action_VALUES(self, id: str, card_name: str, curr_meta_col: ForecastMetaDataSeries, ref_map_id: str, add_blank_row_before: bool = False, add_blank_row_after: bool = False):
        self._add_values_row(ref_map_id = ref_map_id,
                             id = id, 
                             tab_name = card_name, 
                             values = curr_meta_col.get_data_values(), 
                             curr_row_meta_data = curr_meta_col, 
                             row_name = curr_meta_col.get_display_name(),
                             add_blank_row_before = add_blank_row_before,
                             add_blank_row_after = add_blank_row_after)



    # action_DATES
    # BUILD INTERFACE:  handle the DATES action by creating a row of dates
    #  
    # INPUTS:
    #   id = the id of the row
    #   card_name = name of the EXCEL tab to output results
    #   curr_meta_col = ForecastMetaDataSeries meta_data for this column
    # 
    # OUTPUTS:
    #   NA
    def action_DATES(self, id: str, card_name: str, curr_meta_col: ForecastMetaDataSeries, ref_map_id: str, add_blank_row_before: bool = False, add_blank_row_after: bool = False):
        # dates are in strings formatted as: '2027-12-31T00:00:00', so convert to datetime
        # so that openpyxl can correctly pass them to excel
        values = []

        for date in curr_meta_col.get_data_values():
            # if it's string
            if(isinstance(date, str)):
                values.append(datetime.strptime(date, "%Y-%m-%dT%H:%M:%S"))

            # if it's pd.Timestamp
            elif(isinstance(date, pd.Timestamp)):
                values.append(date.to_pydatetime())

            # else throw an error
            else:
                raise ValueError(f"\n*  action_DATES:  invalidate type provided {type(date)}")

        self._add_values_row(ref_map_id = ref_map_id,
                             id = id, 
                             tab_name = card_name, 
                             values = values, 
                             curr_row_meta_data = curr_meta_col, 
                             row_name = curr_meta_col.get_display_name(),
                             add_blank_row_before = add_blank_row_before,
                             add_blank_row_after = add_blank_row_after)
        


    # action_INPUT
    # BUILD INTERFACE:    handle the INPUT action by putting a row of values and blank cells for entering data
    #  
    # INPUTS:
    #   id = the id of the row
    #   card_name = name of the EXCEL tab to output results
    #   curr_meta_col = ForecastMetaDataSeries meta_data for this column
    # 
    # OUTPUTS:
    #   NA
    def action_INPUT(self, id: str, card_name: str, curr_meta_col: ForecastMetaDataSeries, ref_map_id: str, add_blank_row_before: bool = False, add_blank_row_after: bool = False):
        self._add_values_row(ref_map_id = ref_map_id,
                             id = id, 
                             tab_name = card_name, 
                             values = curr_meta_col.get_data_values(), 
                             curr_row_meta_data = curr_meta_col, 
                             row_name = curr_meta_col.get_display_name(),
                             add_blank_row_before = add_blank_row_before,
                             add_blank_row_after = add_blank_row_after)
    


    # action_SUM
    # BUILD INTERFACE:  handle the SUM action by creating a row SUM'd variables based on the PREDs provided
    #  
    # INPUTS:
    #   id = the id of the row
    #   card_name = name of the EXCEL tab to output results
    #   curr_meta_col = ForecastMetaDataSeries meta_data for this column
    # 
    # OUTPUTS:
    #   NA
    def action_SUM(self, id: str, card_name: str, curr_meta_col: ForecastMetaDataSeries, ref_map_id: str, add_blank_row_before: bool = False, add_blank_row_after: bool = False):
        self._add_arith_row(arith_funct = ForecastBuilderExcelArithmeticFunctions.ADD,
                            ref_map_id = ref_map_id,
                            id = id,
                            tab_name = card_name, 
                            values = curr_meta_col.get_data_values(),
                            curr_row_meta_data = curr_meta_col,
                            row_name = curr_meta_col.get_display_name(),
                            add_blank_row_before = add_blank_row_before,
                            add_blank_row_after = add_blank_row_after)




    # action_TOTAL
    # BUILD INTERFACE:  same formula's as total, by may be handled differently given it's a (SUB)TOTAL
    #  
    # INPUTS:
    #   id = the id of the row
    #   card_name = name of the EXCEL tab to output results
    #   curr_meta_col = ForecastMetaDataSeries meta_data for this column
    # 
    # OUTPUTS:
    #   NA
    def action_TOTAL(self, id: str, card_name: str, curr_meta_col: ForecastMetaDataSeries, ref_map_id: str, add_blank_row_before: bool = False, add_blank_row_after: bool = None):
        # if left unspecified, add a blank row after the total line
        if(add_blank_row_after is None):
            add_blank_row_after = True

        self.action_SUM(id = id,
                        card_name = card_name, 
                        curr_meta_col = curr_meta_col, 
                        ref_map_id = ref_map_id, 
                        add_blank_row_before = add_blank_row_before, 
                        add_blank_row_after = add_blank_row_after)




    # action_PROD
    # BUILD INTERFACE:  handle the PROD action by creating a row PROD'd variables based on the PREDs provided
    #  
    # INPUTS:
    #   id = the id of the row
    #   card_name = name of the EXCEL tab to output results
    #   curr_meta_col = ForecastMetaDataSeries meta_data for this column
    # 
    # OUTPUTS:
    #   NA
    def action_PROD(self, id: str, card_name: str, curr_meta_col: ForecastMetaDataSeries, ref_map_id: str, add_blank_row_before: bool = False, add_blank_row_after: bool = False):
        self._add_arith_row(arith_funct = ForecastBuilderExcelArithmeticFunctions.PROD,
                            ref_map_id = ref_map_id,
                            id = id,
                            tab_name = card_name, 
                            values = curr_meta_col.get_data_values(),
                            curr_row_meta_data = curr_meta_col,
                            row_name = curr_meta_col.get_display_name(),
                            add_blank_row_before = add_blank_row_before,
                            add_blank_row_after = add_blank_row_after)




    # action_SUB
    # BUILD INTERFACE:  handle the subtraction action by creating a row of subtracted variables based on the PREDs provided
    #  
    # INPUTS:
    #   id = the id of the row
    #   card_name = name of the EXCEL tab to output results
    #   curr_meta_col = ForecastMetaDataSeries meta_data for this column
    # 
    # OUTPUTS:
    #   NA
    def action_SUB(self, id: str, card_name: str, curr_meta_col: ForecastMetaDataSeries, ref_map_id: str, add_blank_row_before: bool = False, add_blank_row_after: bool = False):
        self._add_arith_row(arith_funct = ForecastBuilderExcelArithmeticFunctions.SUB,
                            ref_map_id = ref_map_id,
                            id = id,
                            tab_name = card_name, 
                            values = curr_meta_col.get_data_values(),
                            curr_row_meta_data = curr_meta_col,
                            row_name = curr_meta_col.get_display_name(),
                            add_blank_row_before = add_blank_row_before,
                            add_blank_row_after = add_blank_row_after)
        



    # action_YEAR_TO_MONTH
    # BUILD INTERFACE:  handle the YEAR_TO_MONTH action by creating a row with the source values expanded to monthly values
    #  
    # INPUTS:
    #   id = the id of the row
    #   card_name = name of the EXCEL tab to output results
    #   curr_meta_col = ForecastMetaDataSeries meta_data for this column
    # 
    # OUTPUTS:
    #   NA
    def action_YEAR_TO_MONTH(self, id: str, card_name: str, curr_meta_col: ForecastMetaDataSeries, ref_map_id: str, add_blank_row_before: bool = False, add_blank_row_after: bool = False):
        self._add_YTM_row(ref_map_id = ref_map_id,
                          id = id,
                          #tab_name = ForecastBuilderExcelRequiredTabs.TIME_CONVERT,
                          tab_name = card_name,
                          values = curr_meta_col.get_data_values(),
                          curr_row_meta_data = curr_meta_col,
                          row_name = curr_meta_col.get_display_name(),
                          add_blank_row_before = add_blank_row_before,
                          add_blank_row_after = add_blank_row_after)



    # action_MONTH_TO_YEAR
    # BUILD INTERFACE:  handle the MONTH_TO_YEAR action by creating a row with the source values collapsed to annual values
    #  
    # INPUTS:
    #   id = the id of the row
    #   card_name = name of the EXCEL tab to output results
    #   curr_meta_col = ForecastMetaDataSeries meta_data for this column
    # 
    # OUTPUTS:
    #   NA
    def action_MONTH_TO_YEAR(self, id: str, card_name: str, curr_meta_col: ForecastMetaDataSeries, ref_map_id: str, add_blank_row_before: bool = False, add_blank_row_after: bool = False):
        self._add_MTY_row(ref_map_id = ref_map_id,
                          id = id,
                          #tab_name = ForecastBuilderExcelRequiredTabs.TIME_CONVERT,
                          tab_name = card_name,
                          values = curr_meta_col.get_data_values(),
                          curr_row_meta_data = curr_meta_col,
                          row_name = curr_meta_col.get_display_name(),
                          add_blank_row_before = add_blank_row_before,
                          add_blank_row_after = add_blank_row_after)



    # action_STEP_INIT
    # BUILD INTERFACE:  handle the STEP_INIT action, this is a dispatcher to a set of internal INIT helper functions
    #  
    # INPUTS:
    #   id = the id of the row
    #   card_name = name of the EXCEL tab to output results
    #   curr_meta_col = ForecastMetaDataSeries meta_data for this column
    # 
    # OUTPUTS:
    #   NA
    def action_STEP_INIT(self, id: str, card_name: str, curr_meta_col: ForecastMetaDataSeries, ref_map_id: str, add_blank_row_before: bool = False, add_blank_row_after: bool = False) -> str:
        step_type = curr_meta_col.get_step_type()

        match step_type:
            case ForecastDataSeriesMetaDataStepTypes.TREATMENT:
                target_card = self._add_step_init_treatment(ref_map_id = ref_map_id,
                                                            id = id,
                                                            tab_name = card_name,
                                                            curr_row_meta_data = curr_meta_col, 
                                                            display_name = curr_meta_col.get_display_name(),
                                                            add_blank_row_before = add_blank_row_before,
                                                            add_blank_row_after = add_blank_row_after)
                
            case ForecastDataSeriesMetaDataStepTypes.SUMMATION:
                target_card = self._add_step_init_summation(ref_map_id = ref_map_id,
                                                            id = id,
                                                            tab_name = card_name,
                                                            curr_row_meta_data = curr_meta_col, 
                                                            display_name = curr_meta_col.get_display_name(),
                                                            add_blank_row_before = add_blank_row_before,
                                                            add_blank_row_after = add_blank_row_after)

            case _:
                target_card = self._add_step_init_default(ref_map_id = ref_map_id,
                                                          id = id, 
                                                          tab_name = card_name, 
                                                          curr_row_meta_data = curr_meta_col, 
                                                          display_name = curr_meta_col.get_display_name(),
                                                          add_blank_row_before = add_blank_row_before,
                                                          add_blank_row_after = add_blank_row_after)
                
        return(target_card)








    # BUILDER INTERNAL FUNCTIONS
    # ==========================

    # _initialize_new_builder
    # All the excel specific set-up steps to start the process of building a new player:
    # copy the template, open a new workbook object, set-up the initial tabs, 
    #  set-up the id to cell references tracking structures, etc.
    #  
    # INPUTS:
    #   TBD
    # 
    # OUTPUTS:
    #   TBD
    def _initialize_new_builder(self, default_num_elements: int):
        # create new workbook or copy template
        if(self.hasTemplate):
            # since openpyxl has no object copy method for a workbook, we have to copy the template workbook
            # to the new filename and then open that
            shutil.copyfile(self.template_location, self.output_location)
            self.player_model = load_workbook(self.output_location)

        else:
            self.player_model = Workbook()

            # if any worksheets were created in the default object, get rid of them
            num_sheets = len(self.player_model.sheetnames)

            if(num_sheets > 0):
                for i in range(num_sheets):
                    ws = ForecastExcelBaseHelpers.get_ws(0, self.player_model)
                    self.player_model.remove(ws)

        # generate the core tabs if they don't exist
        self.row_trackers = {}
        self._generate_core_tabs()


        # initialize core constants / variables

        # setup the EXCEL_FIELD_TO_LETTER_MAP map
        ws = self.player_model[ForecastBuilderExcelRequiredTabs.SUMMARY]

        for element in self.ExcelField:
            self.EXCEL_FIELD_TO_LETTER_MAP[element] = ws.cell(row = 1, column = element.value).column_letter

        # set up the mapper between ForecastDataModel id's and cell references in excel
        self.id_cellref_maps = IdToCellReferenceMaps(default_ref_map_id = self.id, default_num_elements = default_num_elements)

        

    # _finalize_new_builder
    # All the excel specific steps to finalize the new player, mostly saving it to a target file location
    # TODO:  extend this to add save to sharepoint functionality
    #  
    # INPUTS:
    #   TBD
    # 
    # OUTPUTS:
    #   TBD
    def _finalize_new_builder(self):

        # set width of the label column (column 'B' at the time of writing)
        #ws = self.player_model[ForecastBuilderExcelRequiredTabs.SUMMARY]

        for tab in self.row_trackers.keys():
            ws = self.player_model[tab]
            ws.column_dimensions[self.EXCEL_FIELD_TO_LETTER_MAP[self.ExcelField.LABEL]].width = self.LABEL_COL_WIDTH

        # save the file
        self.player_model.save(self.output_location)

        # TODO:  figure out what the return here, the intent is to provide a way to
        # return things in realtime if a builder doesn't save to a file (for example a dynamic JSON file)
        # but have to figure out what (if anything) to return when it does save a file
        return(self.player_model)


    
    # _build_model_excel
    # This kickes off the main dispatcher for the builder, iterates over all the rows in the ForecastMetaDataFrame model, and dispatches
    # different functions the handle the different ACTIONS to build the excel player
    #  
    # INPUTS:
    #   NA
    # 
    # OUTPUTS:
    #   NA
    def _build_model_excel(self):
        self.default_card = ForecastBuilderExcelRequiredTabs.SUMMARY
        self._build_metadataframe_model(meta_data = self.meta_data, default_card = ForecastBuilderExcelRequiredTabs.SUMMARY)
                


    # _build_validation_excel
    # The validation rules dispathcher for the builder, takes the current cell and iterates over all the elements in a list of 
    # validation elements rules, dispatching the different functions the handle the different validation needs of that cell
    #  
    # INPUTS:
    #   curr_cell - openpyxl Cell object, the current cell being working on
    # 
    # OUTPUTS:
    #   NA

    def _build_validation_excel(self,
                                curr_cell_meta_data: ForecastMetaDataSeries, 
                                curr_cell: Cell,
                                data_type: ForecastMetaDataSeriesSchema, 
                                display_type: ForecastMetaDataSeriesSchema, 
                                validation_rules: List[Dict[ForecastDataSeriesMetaDataValidationSchema, Any]],
                                curr_formula: str = None,
                                apply_to_preds: bool = False,
                                list_of_pred_refs: list[str] = None):
        
        if ((apply_to_preds) and (list_of_pred_refs is None)):
            raise ValueError(f"\n*  _build_validation_excel:  apply_to_preds is True, but no pred field in ForecastDataSeries meta_data")
        
        for rule in validation_rules:
            rule_type = list(rule.keys())[0]
            rule_value = rule[rule_type]

            match rule_type:
                case ForecastDataSeriesMetaDataValidationSchema.INPUT_RESTRICTION:
                    self._add_input_restrictions(curr_cell = curr_cell,
                                                 data_type = data_type,
                                                 display_type = display_type,
                                                 restriction = rule_value,
                                                 curr_cell_meta_data = curr_cell_meta_data)

                case ForecastDataSeriesMetaDataValidationSchema.VALUE_CHECK:
                    # if preds exists, they needed to determine where to put the value check, so generate the formulas
                    # since data_validation in excel CANNOT work with Worksheet names (i.e. A1 works, 'Summary'!A1 does not)
                    # we have to remove the worksheet name portion of all the cell references we were given

                    # if apply_to_preds:
                    #     preds = self._ids_to_formula_refs(preds, col_num = curr_cell.column, with_ws_name = False) 

                    # TODO:  Add function that removes tab names from list of ref strings
                    if list_of_pred_refs is not None:
                        list_of_pred_refs = ForecastExcelBaseHelpers.remove_worksheet_names_from_list(list_of_pred_refs)

                    # adds the value comparison
                    self._add_value_comparison(curr_cell = curr_cell,
                                               comparison_type = rule_value,
                                               curr_cell_meta_data = curr_cell_meta_data,
                                               preds = list_of_pred_refs,
                                               curr_formula = curr_formula)

                case _:
                    raise ValueError(f"*  _build_validation_excel:  Unknown validation rule {rule}")




    # BUILDER ACTION HANDLERS
    # =======================

    # _add_values_row
    # Add a row of values to the model
    #  
    # INPUTS:
    #   
    #   tab = the tab to add it to
    #   values = the series of values to add
    #   restrictions = the set of restrictions on data entry
    # 
    # OUTPUTS:
    #   worksheet

    def _add_values_row(self, 
                        ref_map_id: str,
                        id: str,
                        tab_name: str,
                        values: list, 
                        curr_row_meta_data: ForecastMetaDataSeries, 
                        restriction: List[Dict[ForecastDataSeriesMetaDataValidationSchema, Any]] = None, 
                        row_name: str = None, 
                        add_blank_row_before: bool = False, 
                        add_blank_row_after: bool = False):
        
        # if we are adding a blank row before, then add it now
        if(add_blank_row_before):
            self._add_blank_row(tab_name = tab_name)

        # make sure it's a legit tab to go to start writing
        if(tab_name not in self.row_trackers.keys()):
            raise ValueError(f"* add_input_row:  error, requested tab '{tab_name}' not in the list of tabs\n{self.row_trackers.keys()}")
        
        # get meta_data values that we need for this function, if it wasn't provided
        if restriction is None:
            restriction = curr_row_meta_data.get_validation()
        
        # boilerplate set-up of the row
        (curr_cell, data_type, display_type) = self._add_row_setup(ref_map_id, id, tab_name, curr_row_meta_data, row_name)

        
        # create an iterator for the PREDs
        if(curr_row_meta_data.has_preds()):
            # Iterate over all the columns in the row that need values
            num_vals_add = len(values)

            preds_iterator = iter(ForecastPredIterator(col = curr_row_meta_data, address_maps = self.id_cellref_maps, default_card = self.id_cellref_maps.default_ref_map_id, total_elements = num_vals_add))
            list_of_pred_refs = [pred_ref for pred_ref in preds_iterator]
        else:
            list_of_pred_refs = None


        # go cell by cell adding the row values
        for value in values:
            curr_cell.value = value # set the value of the cell
            self._build_validation_excel(curr_cell_meta_data = curr_row_meta_data, curr_cell = curr_cell, data_type = data_type, display_type = display_type, validation_rules = restriction) # add validation rules (if any)
            curr_cell = curr_cell.offset(row = 0, column = 1) # move the current cell pointer by one over to the right of the row
        
        # if we are adding a blank row after, then add it now
        if(add_blank_row_after):
            self._add_blank_row(tab_name = tab_name)
            


    
    # _add_arith_row
    # Add a row of arithmetic formulas (+, -, *, /)
    #  
    # INPUTS:
    #   
    #   tab = the tab to add it to
    #   values = the series of values to run an arithmetic function
    #   type = the data type (for formatting purposes)
    #   restrictions = the set of restrictions on data entry
    # 
    # OUTPUTS:
    #   worksheet

    def _add_arith_row(self, 
                       arith_funct: ForecastBuilderExcelArithmeticFunctions, 
                       ref_map_id: str,
                       id: str, 
                       tab_name: str,
                       values: list,
                       curr_row_meta_data: ForecastMetaDataSeries,
                       preds: list[str] = None, 
                       restriction: List[Dict[ForecastDataSeriesMetaDataValidationSchema, Any]] = None, 
                       row_name: str = None, 
                       add_blank_row_before: bool = False, 
                       add_blank_row_after: bool = False):
        

        # if we are adding a blank row before, then add it now
        if(add_blank_row_before):
            self._add_blank_row(tab_name = tab_name)

        # make sure it's a legit tab to go to start writing
        if(tab_name not in self.row_trackers.keys()):
            raise ValueError(f"* add_arith_row:  error, requested tab '{tab_name}' not in the list of tabs\n{self.row_trackers.keys()}")
        
        # get meta_data values that we need for this function, if it wasn't provided
        if preds is None:
            preds = curr_row_meta_data.get_pred()
        
        if restriction is None:
            restriction = curr_row_meta_data.get_validation()
        
        # boilerplate set-up of the row
        # and place the curr_cell pointer in the row and column to start adding values
        (curr_cell, data_type, display_type)  = self._add_row_setup(ref_map_id, id, tab_name, curr_row_meta_data, row_name)

        # Iterate over all the columns in the row that need values
        num_vals_add = len(values)

        # create an iterator for the PREDs
        preds_iterator = iter(ForecastPredIterator(col = curr_row_meta_data, address_maps = self.id_cellref_maps, default_card = self.id_cellref_maps.default_ref_map_id, total_elements = num_vals_add))


        # generate and write the arith values to the ROW
        for curr_col in range(curr_cell.column, curr_cell.column + num_vals_add):
           list_of_formula_refs = self._ids_to_formula_refs(next(preds_iterator))  # generate a list of formula referenes, and constants
           formula = f"={arith_funct.value.join(list_of_formula_refs)}"
           curr_cell.value = formula # create the formula (i.e. = 'Summary'!A1 + 'Summary'!B1 + 'Summary'!C1)
           
           # add validation rules (if any)
           self._build_validation_excel(curr_cell_meta_data = curr_row_meta_data,
                                        list_of_pred_refs = list_of_formula_refs,
                                        curr_cell = curr_cell, 
                                        data_type = data_type, 
                                        display_type = display_type, 
                                        validation_rules = restriction,
                                        curr_formula = formula,
                                        apply_to_preds = True)
           
           curr_cell = curr_cell.offset(row = 0, column = 1)

        # if we are adding a blank row after, then add it now
        if(add_blank_row_after):
            self._add_blank_row(tab_name = tab_name)
            


    # _add_YTM_row
    # Add a YEAR TO MONTH conversion row
    #  
    # INPUTS:
    #   tab = the tab to add it to
    #   values = the input row
    #   type = the data type (for formatting purposes)
    #   restrictions = the set of restrictions on data entry
    # 
    # OUTPUTS:
    #   worksheet

    def _add_YTM_row(self, 
                     ref_map_id: str,
                     id: str, 
                     tab_name: str,
                     values: list,
                     curr_row_meta_data: ForecastMetaDataSeries,
                     restriction: List[Dict[ForecastDataSeriesMetaDataValidationSchema, Any]] = None, 
                     row_name: str = None, 
                     add_blank_row_before: bool = False, 
                     add_blank_row_after: bool = False):
        
        # if we are adding a blank row before, then add it now
        if(add_blank_row_before):
            self._add_blank_row(tab_name = tab_name)

        # make sure it's a legit tab to go to start writing
        if(tab_name not in self.row_trackers.keys()):
            raise ValueError(f"\n* _add_YTM_row:  error, requested tab '{tab_name}' not in the list of tabs\n{self.row_trackers.keys()}")
        
        # If dates are provided, add the row of monthly dates before the values row
        if curr_row_meta_data.has_arg():
            dates = curr_row_meta_data.get_arg(ForecastDataSeriesMetaDataArgsMTYYTM.DATES)


            if(dates is not None):
                date_row_meta_data = ForecastExcelBaseHelpers.quick_static_date_series(step = curr_row_meta_data.get_step_type(), 
                                                                                       label = "Dates (end-of) (converted from years to months)", 
                                                                                       values = dates)
                self.action_DATES(id = date_row_meta_data.get_id(), card_name = tab_name, curr_meta_col = date_row_meta_data, ref_map_id = ref_map_id, add_blank_row_before=True)
        
        # Iterate over all the columns in the row that need values
        num_vals_add = len(values)

        # create an iterator for the PREDs
        preds_iterator = iter(ForecastPredIterator(col = curr_row_meta_data, address_maps = self.id_cellref_maps, default_card = self.id_cellref_maps.default_ref_map_id, total_elements = num_vals_add))

        if restriction is None:
            restriction = curr_row_meta_data.get_validation()
        
        # boilerplate set-up of the row
        # and place the curr_cell pointer in the row and column to start adding values
        (curr_cell, data_type, display_type)  = self._add_row_setup(ref_map_id, id, tab_name, curr_row_meta_data, row_name)

        i = 0 # counter on the months (if it's % 12 == 0, we've reached a new year)
        curr_pred_col = curr_cell.column # the column for the pred cells (which are the years)

        for curr_col in range(curr_cell.column, curr_cell.column + num_vals_add):
           # if we are an even division of 12, then grab the next year's values from the reference to add
           if i % 12 == 0:
               # formula_ref = self._id_to_formula_ref(preds, col_num = curr_pred_col) # get the value for the current year
               list_of_formula_refs = self._ids_to_formula_refs(next(preds_iterator)) # return a list by default, but will only be one value (or it's an error)
               curr_pred_col += 1 # increment so that next time, we get the following year
               formula = f"={list_of_formula_refs[0]} / 12" # create the formula (i.e. = 'Summary'!A1 / 12)
           
           # add formula and validation rules to the current cell
           # put in the cell value pointing to the formula
           curr_cell.value = formula 
            
           # add validation rules (if any)
           self._build_validation_excel(curr_cell_meta_data = curr_row_meta_data,
                                        list_of_pred_refs = list_of_formula_refs,
                                        curr_cell = curr_cell, 
                                        data_type = data_type, 
                                        display_type = display_type, 
                                        validation_rules = restriction,
                                        curr_formula = formula,
                                        apply_to_preds = True)
            
           curr_cell = curr_cell.offset(row = 0, column = 1)
           i += 1
        
        # if we are adding a blank row after, then add it now
        if(add_blank_row_after):
            self._add_blank_row(tab_name = tab_name)
            
        



    # _add_mty_row
    # Add a MONTH TO YEAR conversion row
    #  
    # INPUTS:
    #   tab = the tab to add it to
    #   values = the input row
    #   type = the data type (for formatting purposes)
    #   restrictions = the set of restrictions on data entry
    # 
    # OUTPUTS:
    #   worksheet

    def _add_MTY_row(self, 
                     ref_map_id: str,
                     id: str, 
                     tab_name: str,
                     values: list,
                     curr_row_meta_data: ForecastMetaDataSeries,
                     #preds: list[str] = None, 
                     restriction: List[Dict[ForecastDataSeriesMetaDataValidationSchema, Any]] = None, 
                     row_name: str = None, 
                     add_blank_row_before: bool = False, 
                     add_blank_row_after: bool = False):
        
        # if we are adding a blank row before, then add it now
        if(add_blank_row_before):
            self._add_blank_row(tab_name = tab_name)

        # make sure it's a legit tab to go to start writing
        if(tab_name not in self.row_trackers.keys()):
            raise ValueError(f"\n* _add_YTM_row:  error, requested tab '{tab_name}' not in the list of tabs\n{self.row_trackers.keys()}")
        
        # If dates are provided, add the row of monthly dates before the values row
        if curr_row_meta_data.has_arg():
            dates = curr_row_meta_data.get_arg("dates")


            if(dates is not None):
                date_row_meta_data = ForecastExcelBaseHelpers.quick_static_date_series(step = curr_row_meta_data.get_step_type(), 
                                                                                       label = "Dates (end-of) (converted from months to years)", 
                                                                                       values = dates)
                self.action_DATES(id = date_row_meta_data.get_id(), card_name = tab_name, curr_meta_col = date_row_meta_data, ref_map_id = ref_map_id)
        
        # Iterate over all the columns in the row that need values
        num_vals_add = len(values*12)

        # create an iterator for the PREDs
        preds_iterator = iter(ForecastPredIterator(col = curr_row_meta_data, address_maps = self.id_cellref_maps, default_card = self.id_cellref_maps.default_ref_map_id, total_elements = num_vals_add))

        if restriction is None:
            restriction = curr_row_meta_data.get_validation()
        # boilerplate set-up of the row
        # and place the curr_cell pointer in the row and column to start adding values
        (curr_cell, data_type, display_type)  = self._add_row_setup(ref_map_id, id, tab_name, curr_row_meta_data, row_name)

        i = 0 # counter on the years (if it's % 12 == 0, we've reached a new year)
        curr_pred_col = curr_cell.column # the column for the pred cells (which are the months)
        list_of_refs_to_sum = [] # list of references to sum for the current year

        #for curr_col in range(curr_cell.column, curr_cell.column + (num_vals_add*12)-1):
        for i in range(num_vals_add):

           # if we are an even division of 12, then grab the next year's values from the reference to add
           #if (curr_col % 12) == 0:
           list_of_refs_to_sum.append(self._ids_to_formula_refs(next(preds_iterator))[0])

           if (i > 0) and ((i+1) % 12 == 0):
               formula_strings = ",".join(list_of_refs_to_sum)
               formula = f"=SUM({formula_strings})" # create the summation formula
               curr_cell.value = formula

               # add validation rules (if any)
               self._build_validation_excel(curr_cell_meta_data = curr_row_meta_data,
                                            list_of_pred_refs = list_of_refs_to_sum,
                                            curr_cell = curr_cell, 
                                            data_type = data_type, 
                                            display_type = display_type, 
                                            validation_rules = restriction,
                                            curr_formula = formula,
                                            apply_to_preds = True)


               curr_cell = curr_cell.offset(row = 0, column = 1) # move the current cell pointer by one over to the right of the row
               list_of_refs_to_sum = [] # reset the list of references to sum for the next year

        # if we are adding a blank row after, then add it now
        if(add_blank_row_after):
            self._add_blank_row(tab_name = tab_name)
            
        




    # BUILDER STEP_INIT HANDLERS
    # ==========================

    #  _add_step_init_default
    # Default handler for STEP_INIT function, adds a blank row and then displays the STEP's display name as a header 3 style
    #  
    # INPUTS:
    #
    #   id - the id the column
    #   tab_name - the target tab_name to place the step information
    #   curr_row_meta_data - the FormatMetaDataSeries information
    #   display_name - the display name to show
    # 
    # OUTPUTS:
    #   worksheet

    def _add_step_init_default(self,
                               ref_map_id: str,
                               id: str,
                               tab_name: str,
                               curr_row_meta_data: ForecastMetaDataSeries, 
                               display_name : str, 
                               add_blank_row_before: bool = False, 
                               add_blank_row_after: bool = False) -> str:
        
        tab_name = self.DEFAULT_CARD

        # if we are adding a blank row before, then add it now
        if(add_blank_row_before):
            self._add_blank_row(tab_name = tab_name)

        # make sure it's a legit tab to go to start writing
        if(tab_name not in self.row_trackers.keys()):
            raise ValueError(f"* add_input_row:  error, requested tab '{tab_name}' not in the list of tabs\n{self.row_trackers.keys()}")

        # add a blank row
        self._add_blank_row(tab_name)

        # boilerplate set-up of the row
        curr_cell  = self._add_row_label_setup(ref_map_id, id, tab_name, curr_row_meta_data)

        curr_cell.value = display_name
        ForecastExcelCellStyleBuilder.generate_init_step_header(curr_cell)

        # if we are adding a blank row after, then add it now
        if(add_blank_row_after):
            self._add_blank_row(tab_name = tab_name)
            
        return(self.DEFAULT_CARD)




    #  _add_step_init_summation
    # Handler for STEP_INIT SUMMATION function, does not add an extra blank line and does not do the row header (looks better in the spreadsheet without it)
    #  
    # INPUTS:
    #
    #   id - the id the column
    #   tab_name - the target tab_name to place the step information
    #   curr_row_meta_data - the FormatMetaDataSeries information
    #   display_name - the display name to show
    # 
    # OUTPUTS:
    #   worksheet

    def _add_step_init_summation(self,
                               ref_map_id: str,
                               id: str,
                               tab_name: str,
                               curr_row_meta_data: ForecastMetaDataSeries, 
                               display_name : str, 
                               add_blank_row_before: bool = False, 
                               add_blank_row_after: bool = False) -> str:
        
        tab_name = self.DEFAULT_CARD

        # # if we are adding a blank row before, then add it now
        # if(add_blank_row_before):
        #     self._add_blank_row(tab_name = tab_name)

        # # make sure it's a legit tab to go to start writing
        # if(tab_name not in self.row_trackers.keys()):
        #     raise ValueError(f"* add_input_row:  error, requested tab '{tab_name}' not in the list of tabs\n{self.row_trackers.keys()}")

        # add a blank row
        self._add_blank_row(tab_name)

        # # boilerplate set-up of the row
        # curr_cell  = self._add_row_label_setup(ref_map_id, id, tab_name, curr_row_meta_data)

        # curr_cell.value = display_name
        # ForecastExcelCellStyleBuilder.generate_init_step_header(curr_cell)

        # # if we are adding a blank row after, then add it now
        # if(add_blank_row_after):
        #     self._add_blank_row(tab_name = tab_name)
            
        return(self.DEFAULT_CARD)




    #  _add_step_init_treatment
    # Handler for STEP_INIT Treatment, creates a new tab for the treatment, displays the treatment details, and links in the appropriate row for the tab it came from
    #
    # INPUTS:
    #
    #   id - the id the column
    #   curr_row_meta_data - the FormatMetaDataSeries information
    #   display_name - the display name to show
    # 
    # OUTPUTS:
    #   worksheet

    def _add_step_init_treatment(self,
                                 ref_map_id: str,
                                 id: str,
                                 tab_name: str,
                                 curr_row_meta_data: ForecastMetaDataSeries, 
                                 display_name : str, 
                                 add_blank_row_before: bool = False, 
                                 add_blank_row_after: bool = False):

        # SETUP
        # ZIV
        # get all the input OBJs
        dict_of_objects = curr_row_meta_data.get_objs()
        treatment_details_model = dict_of_objects.get(ForecastTreatmentStepInitArgs.TREATMENT_TABLE_DATA.value, None)
        treatment_details_meta_data = dict_of_objects.get(ForecastTreatmentStepInitArgs.TREATMENT_TABLE_META_DATA.value, None)
        
        # grab the pre-forecast data if we need it
        if bool(curr_row_meta_data.get_arg(ForecastDataSeriesMetaDataArgsTreatmentStepInit.NEED_PRE_FORECAST_DATA)):
            pre_forecast_inputs_model = dict_of_objects.get(ForecastTreatmentStepInitArgs.PRE_FORECAST_INPUTS_DATA.value, None)
            pre_forecast_inputs_meta_data = dict_of_objects.get(ForecastTreatmentStepInitArgs.PRE_FORECAST_INPUTS_META_DATA.value, None)
            pre_forecast_patient_flow_model = dict_of_objects.get(ForecastTreatmentStepInitArgs.PRE_FORECAST_PATIENT_FLOW_DATA.value, None)
            pre_forecast_patient_flow_meta_data = dict_of_objects.get(ForecastTreatmentStepInitArgs.PRE_FORECAST_PATIENT_FLOW_META_DATA.value, None)

        treatment_tab_name = ""

        # generate a new tab for the treatment
        tab_names = self.player_model.sheetnames
        treatment_tab_name = ForecastExcelBaseHelpers.gen_excel_tab_name(name = display_name, existing_tab_names = tab_names)
        self._create_tab(treatment_tab_name, protect_worksheet = FORECAST_EXCEL_PROTECT_WORKSHEET, workbook = self.player_model, num_sheets = self.NEW_TAB_INSERT_INDEX_FROM_END)

        # if we are adding a blank row before, then add it now
        if(add_blank_row_before):
            self._add_blank_row(tab_name = treatment_tab_name)
            
        # TREATMENT DETAILS SECTION
        self._add_header_row(tab_name = treatment_tab_name, display_name = "Treatment Details")
        self._build_metadataframe_model(meta_data = treatment_details_meta_data, default_card = treatment_tab_name)
        self._add_blank_row(tab_name = treatment_tab_name)

        if bool(curr_row_meta_data.get_arg(ForecastDataSeriesMetaDataArgsTreatmentStepInit.NEED_PRE_FORECAST_DATA)):
            # PRE-FORECAST INPUTS SECTION
            self._add_header_row(tab_name = treatment_tab_name, display_name = "Pre-Forecast Input")
            self._build_metadataframe_model(meta_data = pre_forecast_inputs_meta_data, default_card = treatment_tab_name)
            self._add_blank_row(tab_name = treatment_tab_name)

            # PRE-FORECAST PATIENT FLOW SECTION
            self._add_header_row(tab_name = treatment_tab_name, display_name = "Pre-Forecast # of Patients")
            self._build_metadataframe_model(meta_data = pre_forecast_patient_flow_meta_data, default_card = treatment_tab_name)
            self._add_blank_row(tab_name = treatment_tab_name)

        # if we are adding a blank row after, then add it now
        if(add_blank_row_after):
            self._add_blank_row(tab_name = treatment_tab_name)
            
        return(treatment_tab_name)



    # HELPER FUNCTIONS 
    # ================


    
    
    # TABS

    # _generate_core_tabs
    # Create the initial set of tabs required by the player
    #  
    # INPUTS:
    #   NA
    # 
    # OUTPUTS:
    #   NA

    def _generate_core_tabs(self):
        # get the number of worksheets in workbook
        self.template_num_sheets = len(self.player_model.sheetnames)
        
        # generate the core tabs of the model
        #for tab in self.EXCEL_REQUIRED_WORKBOOK_TABS:
        for tab in ForecastBuilderExcelRequiredTabs:
            self._create_tab(tab.value, protect_worksheet = FORECAST_EXCEL_PROTECT_WORKSHEET, workbook = self.player_model, num_sheets = self.template_num_sheets)



    # _create_tab
    # Adds a new tab tracker and calls the create tab utility
    #  
    # INPUTS:
    #   NA
    # 
    # OUTPUTS:
    #   NA

    def _create_tab(self, tab: str, protect_worksheet: bool, workbook: Workbook, num_sheets: int):
        self.row_trackers[tab] = self.EXCEL_START_ROW    # NOTE:  All row references are 1-based based per openpyxl
        ForecastExcelBaseHelpers.create_ws(tab, protect_worksheet = FORECAST_EXCEL_PROTECT_WORKSHEET, workbook = self.player_model, num_sheets = num_sheets)
        

    

    # VALIDATION & STYLES

    
    # handles setting any formatting and read/write access based on the input restrictions to this cell
    def _add_input_restrictions(self, 
                                curr_cell: Cell, 
                                data_type: ForecastDataSeriesMetaDataDataType, 
                                display_type: ForecastDataSeriesMetaDataDataType,
                                restriction: ForecastDataSeriesMetaDataValidateInputRestrictions,
                                curr_cell_meta_data: ForecastMetaDataSeries = None):

        # if token check is requested, the figure out if the current value matches the ForecastDataModel approved editable values token
        # if it does, make it read_write, otherwise make it read only
        if(restriction == ForecastDataSeriesMetaDataValidateInputRestrictions.TOKEN_CHECK):
            if(curr_cell.value == ForecastDataModel.EDITABLE_VALUES_TOKEN):
                restriction = ForecastDataSeriesMetaDataValidateInputRestrictions.READ_WRITE
            else:
                restriction = ForecastDataSeriesMetaDataValidateInputRestrictions.READ_ONLY

        # # based on the restriction and the data type, apply cell protection and the appropriate style
        ForecastExcelCellStyleBuilder.generate_cell_data_type_style(curr_cell = curr_cell, 
                                                                    display_type = display_type, 
                                                                    restriction = restriction,
                                                                    curr_cell_meta_data = curr_cell_meta_data)
        
        ForecastExcelValidationRuleBuilder.generate_data_entry_rule(curr_cell = curr_cell, 
                                                                    data_type = data_type,
                                                                    error_title = "Invalid entry",
                                                                    prompt = None,
                                                                    prompt_title = None,
                                                                    allow_blank = True,
                                                                    curr_cell_meta_data = curr_cell_meta_data)



    # handles a validation involving a comparison of an value to another value
    def _add_value_comparison(self,
                              curr_cell: Cell,
                              comparison_type: ForecastDataSeriesMetaDataComparisonType,
                              curr_cell_meta_data: ForecastMetaDataSeries,
                              preds : list[str] = None,
                              curr_formula: str = None):
            
            if(comparison_type not in [ForecastDataSeriesMetaDataComparisonType.BETWEEN, ForecastDataSeriesMetaDataComparisonType.NOT_BETWEEN]):
                # get the limit value for the comparison check
                comparison_id_or_const = curr_cell_meta_data.get_arg(comparison_type)
                #comparison_id_or_const = curr_cell_meta_data.meta_data[ForecastMetaDataSeriesSchema.ARGS][comparison_type]

                ForecastExcelValidationRuleBuilder.generate_comparison_rule(curr_cell = curr_cell,
                                                                            comparison_type = comparison_type,
                                                                            value = comparison_id_or_const,
                                                                            preds = preds,
                                                                            error_title = "Invalid entry",
                                                                            prompt = None,
                                                                            prompt_title = None,
                                                                            allow_blank = True,
                                                                            curr_formula = curr_formula,
                                                                            curr_cell_meta_data = curr_cell_meta_data)
            else:
                # get the limit value for the comparison check
                min_comparison_id_or_const = curr_cell_meta_data.get_arg(comparison_type)[0]
                max_comparison_id_or_const = curr_cell_meta_data.get_arg(comparison_type)[1]

                ForecastExcelValidationRuleBuilder.generate_comparison_rule(curr_cell = curr_cell,
                                                            comparison_type = comparison_type,
                                                            value = min_comparison_id_or_const,
                                                            preds = preds,
                                                            error_title = "Invalid entry",
                                                            prompt = None,
                                                            prompt_title = None,
                                                            allow_blank = True,
                                                            value2 = max_comparison_id_or_const,
                                                            curr_formula = curr_formula,
                                                            curr_cell_meta_data = curr_cell_meta_data)




    

    # ROW


    # _add_row_setup
    # Common boilerplate for all _add actions (values, arith function) to add a new row for an id
    # INPUT
    #   id - ForecastMetaDataSeries ID
    #   tab_name - Tab to add the row to
    #   curr_row_meta_data - ForecastMetaDataSeries information to grab additional row information
    #   row_name (optional) - if a row name (label) should be added, put it here
    #
    # OUTPUT:
    #   curr_cell - a reference to an excel spreadsheet pointing to the first location to write a value
    #   data_type - the data_type to be written
    #   display_type - the data_type to be displayed (for example, we may have a float, but needs to be displayed as an int)
    def _add_row_setup(self, 
                       ref_map_id: str,
                       id: str, 
                       tab_name: str, 
                       curr_row_meta_data: ForecastMetaDataSeries, 
                       row_name: str = None) -> Tuple[Any, ForecastDataSeriesMetaDataDataType, ForecastDataSeriesMetaDataDataType]:

        # grab meta-data that we always grab and return it so we don't have to do this in every function call
        data_type = curr_row_meta_data.get_data_type()
        display_type = curr_row_meta_data.get_display_type()

        # set-up for the adding a row, get a worksheet for the tab name, and the next row to add to
        ws = self.player_model[tab_name]
        curr_row = self.row_trackers[tab_name]
        start_row_ref = ws.cell(row = curr_row, column = ForecastBuilderExcelTB.ExcelField.START)
        
        # add any common row prefix stuff if needed:
        # row label
        if(row_name is not None):
            curr_cell = ws.cell(row = curr_row, column = ForecastBuilderExcelTB.ExcelField.LABEL)
            self._add_label(value = row_name, curr_cell = curr_cell)

        # save the start cell ref for this id, and the tab_name (tab name not needed, but for convenience)
        if(curr_row_meta_data.has_data_values()):
            self._track_add_row(ref_map_id = ref_map_id, id = id, tab_name = tab_name, cell_ref = start_row_ref, num_elements = len(curr_row_meta_data.get_data_values()))
        else:
            self._track_add_row(ref_map_id = ref_map_id, id = id, tab_name = tab_name, cell_ref = start_row_ref, num_elements = 0)

        # move current cell to the start of the values entry
        curr_cell = ws.cell(row = curr_row, column = ForecastBuilderExcelTB.ExcelField.VALUES)

        return(curr_cell, data_type, display_type)
    


    # _add_row_label_setup
    # Common boilerplate for adding a label not inputs or calculated values
    # INPUT
    #   id - ForecastMetaDataSeries ID
    #   tab_name - Tab to add the row to
    #   curr_row_meta_data - ForecastMetaDataSeries information to grab additional row information
    #   row_name (optional) - if a row name (label) should be added, put it here
    #
    # OUTPUT:
    #   curr_cell - a reference to an excel spreadsheet pointing to the first location to write a value
    #   data_type - the data_type to be written
    #   display_type - the data_type to be displayed (for example, we may have a float, but needs to be displayed as an int)

    def _add_row_label_setup(self, ref_map_id: str, id: str, tab_name: str, curr_row_meta_data: ForecastMetaDataSeries, column = ExcelField.START):
        ws = self.player_model[tab_name]
        curr_row = self.row_trackers[tab_name]
        curr_cell = ws.cell(row = curr_row, column = column)
        
        # save the start cell ref for this id, and the tab_name (tab name not needed, but for convenience)
        if(curr_row_meta_data.has_data_values()):
            self._track_add_row(ref_map_id = ref_map_id, id = id, tab_name = tab_name, cell_ref = curr_cell, num_elements = len(curr_row_meta_data.get_data_values()))
        else:
            self._track_add_row(ref_map_id = ref_map_id, id = id, tab_name = tab_name, cell_ref = curr_cell, num_elements = 0)

        # # save the start cell ref for this id, and the tab_name (tab name not needed, but for convenience)
        # self._track_add_row(ref_map_id = ref_map_id, id = id, tab_name = tab_name, cell_ref = curr_cell) # store this 

        return(curr_cell)


    # _add_blank_row
    # Add a blank row without any ids
    #
    # INPUT
    #   tab_name - Tab to add the row to
    #
    # OUTPUT:
    # NA
    def _add_blank_row(self, tab_name: str):
        self.row_trackers[tab_name] += 1

    # _add_header_row
    # Add a header row without any ids
    #
    # INPUT
    #   tab_name - Tab to add the row to
    #   display_name - The name for the header
    #
    # OUTPUT:
    # NA
    # ZIV
    def _add_header_row(self, tab_name: str, display_name: str, column = ExcelField.START):
        ws = self.player_model[tab_name]
        curr_row = self.row_trackers[tab_name]
        curr_cell = ws.cell(row = curr_row, column = column)

        curr_cell.value = display_name
        ForecastExcelCellStyleBuilder.generate_init_step_header(curr_cell)
        
        self.row_trackers[tab_name] += 1




    # _track_add_row
    # handles the boilerplate bookkeeping whenever a new row is added to the model
    #
    # INPUTS:
    #   id = ForecastMetaDataSeries id for the row
    #   tab_name = tab name for the row (mostly for convenience so we don't have to look it up in cell reference)
    #   cell_ref = (cell) cell object for the start of the row
    #
    # OUTPUTS:
    #

    def _track_add_row(self, ref_map_id: str, id: str, tab_name: str, cell_ref: Cell, num_elements: int, store_ref = True):
        if(store_ref):
             self.id_cellref_maps.add(ref_map_id = ref_map_id, id = id, tab_name = tab_name, cell_ref = cell_ref, num_elements = num_elements)  # add the id of the row to our mapper

        self.row_trackers[tab_name] += 1  # increment our row tracker for this tab, so we know location of next row to add




    # FORMULAS

    # _ids_to_formula_refs
    # given a list of either constants (int or float) or ForecastMetaData IDs (strings)
    # and a COLUMN in the spreadsheet, generates a string reference suitable for inclusion in a formula (i.e. 'Summary'!A1)
    # this is useful when you need to generate all the formula values in the same column, but different rows for a forumla
    # i.e. ('Summary'!A1 + 'Summary'!B1 + 'Summary'!C1)
    # NOTE:  the list HAS TO BE A COLUMN... because only one number is provided for the column offset, so if they aren't all lined up
    #        in the same column, the formula will be wrong.
    # 
    # INPUTS:
    #   list_of_ids - list with a mix of ForecastMetaDataSeries IDs or constants (ints or floats)
    #
    # OUTPUTS:
    #   list of strings - each string being either a constant (string version of float or int) or a cell reference suitable for an excel formula

    # def _ids_to_formula_refs(self, list_of_ids: list[int | str | float], col_num: int, with_ws_name: bool = True) -> list[str]:
    #     list_of_refs = []

    #     # iterate of all the list
    #     for id_or_const in list_of_ids:
    #         list_of_refs.append(self._id_to_formula_ref(id_or_const, col_num, with_ws_name = with_ws_name))

    #     return(list_of_refs)
    
    def _ids_to_formula_refs(self, list_of_ids: dict, with_ws_name: bool = True) -> list[str]:
        list_of_refs = []

        # iterate of all the list
        for id in list_of_ids.keys():
            list_of_refs.append(self._id_to_formula_ref(list_of_ids[id], with_ws_name = with_ws_name))

        return(list_of_refs)
    

    # _id_to_formula_ref
    # given a value which is either constants (int or float) or ForecastMetaData IDs (strings)
    # and a column in the spreadsheet, generates a string reference suitable for inclusion in a formula (i.e. 'Summary'!A1)
    # this is useful when you need to generate all the formula values in the same column, but different rows for a forumla
    # i.e. ('Summary'!A1 + 'Summary'!B1 + 'Summary'!C1)
    # 
    # INPUTS:
    #   id - (list) get the information for a SINGLE ID as a list with three values:
    #       tab_name = tab name for the id
    #       constant or Cell = either an int/float number, or an openpyxl Cell object holding the reference to the cell
    #       num_elements - (int) the number of elements in the row
    #   with_ws_name - (optional)(bool) when creating the cell formula reference, keep the worksheet name (i.e. 'Summary'!A1) or not (i.e. A1)
    #   
    #
    # OUTPUTS:
    #   cell_formula_reference - (str) either a constant (string version of float or int) or a cell reference suitable for an excel formula            

    def _id_to_formula_ref(self, id: list, with_ws_name: bool = True) -> str:
            tab_name = None
            cell_or_const = None
            
            # if this is a list or tuple, then the id includes the tab name as well as the cell reference, so split them out
            # ZIV
            if(isinstance(id, list | tuple)):
               tab_name = id[0]
               cell_or_const = id[1]
            else:
                cell_or_const = id

            # determine if this is an ID or a constant
            # if constant, return a string version of the constant
            if isinstance(cell_or_const, int | float):
                return(str(cell_or_const))
            
            # get the addess for this cell
            elif isinstance(cell_or_const, Cell):
                return(ForecastExcelBaseHelpers.cell_to_formula_ref(cell_or_const, with_ws_name=with_ws_name))           

            else:
                raise ValueError(f"\n * _id_to_formula_ref:  Error id '{cell_or_const}' has invalid type '{type(cell_or_const)}'.")

    # CELL

    # _add_label
    # Add a label as the cell value/style
    # INPUT
    #   value = value of the label
    #   cell = openpyxl Cell reference (object)
    #
    # OUTPUT:
    #   NA

    def _add_label(self, value: str, curr_cell: Cell):
        curr_cell.value = value
        ForecastExcelCellStyleBuilder.generate_row_header_label(curr_cell)
        curr_cell.protection = Protection(locked = True, hidden = False)

    
